# website/views.py

import logging
import os
import re
import json
import html
from pathlib import Path

from all_roads.models import (
    Segment,
    Route,
    Road,
    Address,
    State,
    SubSegment,
    Defect,
    DefectType,
    RootCauseAnalysis,
    RootCauseDetail,
    PhysicalInspection,
    PhysicalInspectionAnalysis,
    PhysicalInspectionCharacteristic,
    Library,
)
from all_roads.services import refresh_segment_and_subsegments, refresh_subsegments_from_google
from all_roads.tasks import refresh_segments_task
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from django.contrib.auth import get_user_model
from django.core.paginator import Paginator
from django.db import models, transaction
from django.db.models.deletion import ProtectedError
from django.db.models import Sum, Count, Q, IntegerField, Max
from django.db.models.functions import Cast, Trim, Upper
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils.timesince import timesince
from django.utils import timezone
from .forms import UploadSegmentsForm, UploadSubSegmentsForm
import csv
import io
from django.utils.safestring import mark_safe

logger = logging.getLogger(__name__)

# ---- Status buckets used for counts/mini-chart (hex codes) ----
STATUS_BUCKETS = {
    "good": {"codes": ["05700B"]},                   # Good (>=80 km/h)
    "tolerable": {"codes": ["00CC00"]},              # Tolerable (70 to <80 km/h)
    "intolerable": {"codes": ["FF9966"]},            # Intolerable (60 to <70 km/h)
    "failed": {"codes": ["FF5050"]},                 # Failed (<60 km/h)
    "no_response": {"codes": ["666699"]},            # Unknown / no response
}

# ---- Pagination options (Point 5) ----
PAGE_SIZE_DEFAULT = 25
PAGE_SIZE_OPTIONS = [25, 50, 100]

DOCS_DIR = Path(__file__).resolve().parent.parent / "docs"
DOC_FILENAME_LABELS = {
    "INDEX.md": "Index",
    "PROJECT_STATUS_REPORT.md": "Project Status Report",
    "FEATURES.md": "Features",
    "DEPLOYMENT.md": "Deployment",
    "ARCHITECTURE.md": "Architecture",
    "API.md": "API",
    "OPERATIONS.md": "Operations",
    "styleguide-ferma-platform.md": "Styleguide",
}
DOC_FILENAME_ORDER = [
    "INDEX.md",
    "PROJECT_STATUS_REPORT.md",
    "FEATURES.md",
    "DEPLOYMENT.md",
    "ARCHITECTURE.md",
    "API.md",
    "OPERATIONS.md",
    "styleguide-ferma-platform.md",
]


def _library_file_type_from_name(filename):
    ext = os.path.splitext((filename or "").lower())[1]
    if ext in {".doc", ".docx", ".txt", ".rtf", ".odt"}:
        return Library.FILE_TYPE_DOCUMENT
    if ext in {".xls", ".xlsx", ".ods"}:
        return Library.FILE_TYPE_SPREADSHEET
    if ext == ".pdf":
        return Library.FILE_TYPE_PDF
    if ext == ".csv":
        return Library.FILE_TYPE_CSV
    if ext in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}:
        return Library.FILE_TYPE_IMAGE
    if ext in {".ppt", ".pptx", ".odp"}:
        return Library.FILE_TYPE_PRESENTATION
    if ext in {".geojson", ".json", ".kml", ".kmz", ".shp"}:
        return Library.FILE_TYPE_GEO_DATA
    return Library.FILE_TYPE_OTHER


def _slugify_doc_name(value):
    return re.sub(r"[^a-z0-9]+", "-", str(value or "").strip().lower()).strip("-")


def _render_doc_inline(text):
    escaped = html.escape(text or "")
    escaped = re.sub(r"`([^`]+)`", lambda m: f"<code>{m.group(1)}</code>", escaped)
    escaped = re.sub(
        r"\[([^\]]+)\]\(([^)]+)\)",
        lambda m: f'<a href="{html.escape(m.group(2), quote=True)}" target="_blank" rel="noopener noreferrer">{m.group(1)}</a>',
        escaped,
    )
    return escaped


def _render_markdown_without_dependency(content):
    lines = (content or "").splitlines()
    parts = []
    in_code_block = False
    code_lines = []
    list_type = None

    def close_list():
        nonlocal list_type
        if list_type:
            parts.append(f"</{list_type}>")
            list_type = None

    def close_code_block():
        nonlocal in_code_block, code_lines
        if in_code_block:
            code_html = html.escape("\n".join(code_lines))
            parts.append(f"<pre><code>{code_html}</code></pre>")
            in_code_block = False
            code_lines = []

    for raw_line in lines:
        line = raw_line.rstrip("\n")
        stripped = line.strip()

        if stripped.startswith("```"):
            close_list()
            if in_code_block:
                close_code_block()
            else:
                in_code_block = True
                code_lines = []
            continue

        if in_code_block:
            code_lines.append(line)
            continue

        if not stripped:
            close_list()
            continue

        heading_match = re.match(r"^(#{1,6})\s+(.*)$", stripped)
        if heading_match:
            close_list()
            level = len(heading_match.group(1))
            parts.append(f"<h{level}>{_render_doc_inline(heading_match.group(2))}</h{level}>")
            continue

        blockquote_match = re.match(r"^>\s?(.*)$", stripped)
        if blockquote_match:
            close_list()
            parts.append(f"<blockquote><p>{_render_doc_inline(blockquote_match.group(1))}</p></blockquote>")
            continue

        unordered_match = re.match(r"^\s*-\s+(.*)$", line)
        if unordered_match:
            if list_type != "ul":
                close_list()
                parts.append("<ul>")
                list_type = "ul"
            parts.append(f"<li>{_render_doc_inline(unordered_match.group(1))}</li>")
            continue

        ordered_match = re.match(r"^\s*\d+\.\s+(.*)$", line)
        if ordered_match:
            if list_type != "ol":
                close_list()
                parts.append("<ol>")
                list_type = "ol"
            parts.append(f"<li>{_render_doc_inline(ordered_match.group(1))}</li>")
            continue

        close_list()
        parts.append(f"<p>{_render_doc_inline(stripped)}</p>")

    close_list()
    close_code_block()
    return "".join(parts)


def _load_library_documentation(selected_slug):
    items = []
    ordered_files = []
    for filename in DOC_FILENAME_ORDER:
        path = DOCS_DIR / filename
        if path.exists() and path.suffix.lower() == ".md":
            ordered_files.append(path)

    for path in sorted(DOCS_DIR.glob("*.md")):
        if path not in ordered_files:
            ordered_files.append(path)

    for path in ordered_files:
        label = DOC_FILENAME_LABELS.get(path.name, path.stem.replace("_", " ").replace("-", " ").title())
        slug = _slugify_doc_name(path.stem)
        items.append(
            {
                "slug": slug,
                "label": label,
                "path": path,
            }
        )

    selected_item = None
    if selected_slug:
        selected_item = next((item for item in items if item["slug"] == selected_slug), None)

    if selected_item is None and items:
        selected_slug = ""

    rendered_html = ""
    if selected_item:
        try:
            rendered_html = mark_safe(
                _render_markdown_without_dependency(selected_item["path"].read_text(encoding="utf-8"))
            )
        except Exception as exc:
            rendered_html = mark_safe(f"<p>Could not load documentation. ({exc})</p>")

    return {
        "documentation_items": [{"slug": item["slug"], "label": item["label"]} for item in items],
        "selected_documentation_slug": selected_item["slug"] if selected_item else "",
        "selected_documentation_label": selected_item["label"] if selected_item else "",
        "selected_documentation_html": rendered_html,
    }


def _get_assumed_project_user():
    User = get_user_model()
    existing = (
        User.objects.filter(first_name__iexact="Amina", last_name__iexact="Bello")
        .order_by("id")
        .first()
    )
    if existing:
        return existing

    base_username = "amina.bello"
    username = base_username
    suffix = 1
    while User.objects.filter(username=username).exists():
        suffix += 1
        username = f"{base_username}{suffix}"

    create_kwargs = {"username": username}
    if hasattr(User, "EMAIL_FIELD") and User.EMAIL_FIELD:
        email_field = User.EMAIL_FIELD
        create_kwargs[email_field] = f"{username}@example.com"

    manager = User.objects
    if hasattr(manager, "create_user"):
        user = manager.create_user(password=None, **create_kwargs)
    else:
        user = manager.create(**create_kwargs)
    user.first_name = "Amina"
    user.last_name = "Bello"
    if hasattr(user, "set_unusable_password"):
        user.set_unusable_password()
    user.save()
    return user


TERMINAL_DEFECT_STATUSES = {
    Defect.WORKFLOW_REPAIR_ONGOING,
    Defect.WORKFLOW_REPAIR_COMPLETE,
}


def _get_latest_defect(subsegment):
    return (
        Defect.objects.filter(subsegment=subsegment)
        .order_by("-modified", "-id")
        .first()
    )


def _defect_condition_from_subsegment(subsegment):
    status_code = (getattr(subsegment, "status", "") or "").replace("#", "").upper()
    if status_code == "FF5050":
        return Defect.CONDITION_FAILED
    if status_code == "FF9966":
        return Defect.CONDITION_INTOLERABLE
    if status_code in {"00CC00", "05700B"}:
        return Defect.CONDITION_TOLERABLE
    return Defect.CONDITION_BAD


def _subsegment_allows_defect_creation(subsegment):
    status_code = (getattr(subsegment, "status", "") or "").replace("#", "").upper()
    return status_code != "666699"


def _get_or_create_active_defect(subsegment):
    if not _subsegment_allows_defect_creation(subsegment):
        raise ValueError("Cannot create a defect for a sub-segment with No response status.")
    latest = _get_latest_defect(subsegment)
    if latest and latest.workflow_status not in TERMINAL_DEFECT_STATUSES:
        return latest, False
    defect = Defect.objects.create(
        subsegment=subsegment,
        workflow_status=Defect.WORKFLOW_DRAFT,
        condition=_defect_condition_from_subsegment(subsegment),
    )
    return defect, True


def _update_defect_status_from_rca(defect, rca_status):
    if not defect or defect.workflow_status in TERMINAL_DEFECT_STATUSES:
        return
    if rca_status == RootCauseAnalysis.STATUS_COMPLETE:
        next_status = Defect.WORKFLOW_RCA
    else:
        next_status = Defect.WORKFLOW_DRAFT
    if defect.workflow_status != next_status:
        defect.workflow_status = next_status
        defect.save(update_fields=["workflow_status", "modified"])


def _sync_defect_status_from_physical(defect, physical_status):
    if not defect or defect.workflow_status in TERMINAL_DEFECT_STATUSES:
        return
    if physical_status == PhysicalInspection.STATUS_COMPLETE:
        next_status = Defect.WORKFLOW_PHYSICAL
    else:
        next_status = Defect.WORKFLOW_RCA
    if defect.workflow_status != next_status:
        defect.workflow_status = next_status
        defect.save(update_fields=["workflow_status", "modified"])


def _ensure_physical_draft_for_defect(defect):
    if not defect:
        return None
    inspection = (
        PhysicalInspection.objects.filter(defect=defect)
        .order_by("-updated_at", "-id")
        .first()
    )
    if inspection:
        return inspection
    return PhysicalInspection.objects.create(
        subsegment=defect.subsegment,
        defect=defect,
        status=PhysicalInspection.STATUS_DRAFT,
    )


def _touch_defect_modified(defect):
    if not defect:
        return
    defect.modified = timezone.now()
    defect.save(update_fields=["modified"])


def _resolve_origin_back_link(origin, default_label, default_href):
    origin_map = {
        "approvals": ("Approvals", reverse("library_approvals")),
        "root_cause": ("Root Cause Analysis", reverse("engineering_admin")),
        "physical": ("Physical Inspection", reverse("physical_inspection")),
        "solution": ("Solution Design", reverse("library_solution_design")),
        "history": ("Archive", reverse("library_history")),
        "archive": ("Archive", reverse("library_history")),
        "overview": ("Overview", reverse("engineering_admin_overview")),
    }
    label, href = origin_map.get(origin, (default_label, default_href))
    return f"Back to {label}", href

def landing(request):
    return render(request, "website/landing.html", {"active_page": "home"})


def _overview_metrics(qs):
    agg = qs.aggregate(total_length=Sum("distance"), total_segments=Count("id"))
    return {
        "total_length": (agg["total_length"] or Decimal("0.00")).quantize(Decimal("0.01")),
        "total_segments": agg["total_segments"] or 0,
        "counts": {k: qs.filter(status__in=v["codes"]).count() for k, v in STATUS_BUCKETS.items()},
    }


def _format_fixed_coord(value, places=5):
    try:
        if value is None:
            return "-"
        return f"{Decimal(value):.{places}f}"
    except (InvalidOperation, ValueError, TypeError):
        return "-"


def _format_segment_point_display(name, lat, lon):
    point_name = (name or "").strip() or "-"
    return f"{point_name} ({_format_fixed_coord(lat)}, {_format_fixed_coord(lon)})"


def _format_km_total(value):
    if value is None:
        return ""
    try:
        return f"{Decimal(value).quantize(Decimal('0.01')):,.2f} km"
    except (InvalidOperation, ValueError, TypeError):
        return ""


def _build_inventory_context(request, active_page="inventory"):
    qs = Segment.objects.select_related("route", "start_point", "end_point").all()
    selected_road = request.GET.get("road") or ""
    selected_route = request.GET.get("route") or ""
    selected_state = (request.GET.get("state") or "").strip()

    # Keep one active filter at a time.
    if selected_road:
        selected_route = ""
        selected_state = ""
        qs = qs.filter(route__road__road=selected_road)
    elif selected_route:
        selected_road = ""
        selected_state = ""
        qs = qs.filter(route__route=selected_route)
    elif selected_state:
        selected_road = ""
        selected_route = ""
        qs = qs.filter(state__iexact=selected_state)

    roads = Road.objects.only("road").order_by("road")
    routes = Route.objects.only("route").order_by("route")
    states = State.objects.only("state").order_by("state").values_list("state", flat=True)

    filters = {}
    if selected_road:
        filters["road"] = selected_road
    if selected_route:
        filters["route"] = selected_route
    if selected_state:
        filters["state"] = selected_state

    metrics = _overview_metrics(qs)
    all_rows = list(qs.order_by("route__route", "code")[:12])
    focus_segment = all_rows[0] if all_rows else None
    unique_route_count = qs.values("route_id").distinct().count()
    selected_route_obj = Route.objects.filter(route=selected_route).only("route", "details").first() if selected_route else None
    selected_route_segment_count = (
        Segment.objects.filter(route__route=selected_route).count() if selected_route else None
    )
    selected_route_segments = (
        Segment.objects.select_related("route", "start_point", "end_point")
        .filter(route__route=selected_route)
        .order_by("index", "code")
        if selected_route
        else Segment.objects.none()
    )
    first_segment = selected_route_segments.first() if selected_route else None
    last_segment = selected_route_segments.last() if selected_route else None
    summary_start_point = ""
    summary_end_point = ""
    if first_segment:
        start_name = (
            first_segment.start_point.name
            if first_segment.start_point_id and first_segment.start_point and first_segment.start_point.name
            else (first_segment.name or first_segment.state or "")
        )
        summary_start_point = _format_segment_point_display(
            start_name,
            first_segment.start_lat,
            first_segment.start_lon,
        )
    if last_segment:
        end_name = (
            last_segment.end_point.name
            if last_segment.end_point_id and last_segment.end_point and last_segment.end_point.name
            else (last_segment.name or last_segment.state or "")
        )
        summary_end_point = _format_segment_point_display(
            end_name,
            last_segment.end_lat,
            last_segment.end_lon,
        )
    selected_route_total_length = (
        selected_route_segments.aggregate(total_length=Sum("distance")).get("total_length")
        if selected_route
        else None
    )

    segment_summary = {
        "route": selected_route_obj.route if selected_route_obj else "",
        "length": _format_km_total(selected_route_total_length),
        "start_point": summary_start_point,
        "end_point": summary_end_point,
        "passes_through": (selected_route_obj.details or "") if selected_route_obj else "",
        "number_of_segments": selected_route_segment_count if selected_route_obj else "",
    }

    return {
        "active_page": active_page,
        "segments": qs.order_by("route__route", "index", "code")[:50],
        "roads": roads,
        "routes": routes,
        "states": list(states),
        "selected_road": selected_road,
        "selected_route": selected_route,
        "selected_state": selected_state,
        "filters_qs": urlencode(filters),
        "number_routes": unique_route_count,
        "segment_length_total": "----",
        "focus_segment": focus_segment,
        "report_rows": all_rows[:3],
        "segment_summary": segment_summary,
        "show_segment_summary": bool(selected_route_obj),
        "selected_route_segments": selected_route_segments,
        **metrics,
    }


def road_inventory(request):
    context = _build_inventory_context(request, active_page="inventory")
    return render(request, "website/road_inventory.html", context)


def _build_road_condition_context(request):
    qs = Segment.objects.select_related("route", "start_point", "end_point").all()
    current_view = request.GET.get("view") or "map"
    selected_road = request.GET.get("road") or ""
    selected_route = request.GET.get("route") or ""
    selected_state = (request.GET.get("state") or "").strip()
    selected_speed = (request.GET.get("speed") or "").strip().lower()

    # Keep one active filter at a time.
    if selected_road:
        selected_route = ""
        selected_state = ""
        selected_speed = ""
        qs = qs.filter(route__road__road=selected_road)
    elif selected_route:
        selected_road = ""
        selected_state = ""
        selected_speed = ""
        qs = qs.filter(route__route=selected_route)
    elif selected_state:
        selected_road = ""
        selected_route = ""
        selected_speed = ""
        qs = qs.filter(state__iexact=selected_state)
    elif selected_speed in STATUS_BUCKETS:
        selected_road = ""
        selected_route = ""
        selected_state = ""
        qs = qs.filter(status__in=STATUS_BUCKETS[selected_speed]["codes"])

    roads = Road.objects.only("road").order_by("road")
    routes = Route.objects.only("route").order_by("route")
    states = State.objects.only("state").order_by("state").values_list("state", flat=True)
    speed_options = [
        ("good", "Good"),
        ("tolerable", "Tolerable"),
        ("intolerable", "Intolerable"),
        ("failed", "Failed"),
        ("no_response", "No response"),
    ]

    paginator = Paginator(qs.order_by("route__route", "index", "code"), 50)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    filters = {}
    if current_view:
        filters["view"] = current_view
    if selected_road:
        filters["road"] = selected_road
    if selected_route:
        filters["route"] = selected_route
    if selected_state:
        filters["state"] = selected_state
    if selected_speed:
        filters["speed"] = selected_speed

    metrics = _overview_metrics(qs)
    all_rows = list(qs.order_by("route__route", "code")[:12])

    return {
        "active_page": "road_condition",
        "segments": page_obj.object_list,
        "page_obj": page_obj,
        "roads": roads,
        "routes": routes,
        "states": list(states),
        "speed_options": speed_options,
        "selected_road": selected_road,
        "selected_route": selected_route,
        "selected_state": selected_state,
        "selected_speed": selected_speed,
        "current_view": current_view,
        "filters_qs": urlencode(filters),
        "report_rows": all_rows[:3],
        **metrics,
    }


def road_condition(request):
    context = _build_road_condition_context(request)
    return render(request, "website/road_condition.html", context)


def road_condition_save_draft(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "POST method required."}, status=405)

    subsegment_codes = [
        (code or "").strip()
        for code in request.POST.getlist("subsegment_codes[]")
        if (code or "").strip()
    ]
    if not subsegment_codes:
        return JsonResponse(
            {"ok": True, "created_count": 0, "blocked_codes": [], "not_found_codes": [], "message": "No selected sub-segments."}
        )

    created_count = 0
    blocked_codes = []
    not_found_codes = []

    with transaction.atomic():
        for code in subsegment_codes:
            subsegment = SubSegment.objects.filter(code__iexact=code).first()
            if subsegment is None:
                not_found_codes.append(code)
                continue
            if not _subsegment_allows_defect_creation(subsegment):
                blocked_codes.append(subsegment.code)
                continue

            latest_defect = _get_latest_defect(subsegment)
            if latest_defect and latest_defect.workflow_status not in TERMINAL_DEFECT_STATUSES:
                blocked_codes.append(subsegment.code)
                continue

            defect = Defect.objects.create(
                subsegment=subsegment,
                workflow_status=Defect.WORKFLOW_DRAFT,
                condition=_defect_condition_from_subsegment(subsegment),
                engineer=request.user if request.user.is_authenticated else None,
            )
            RootCauseAnalysis.objects.create(
                subsegment=subsegment,
                defect=defect,
                location=(subsegment.code or "Unknown")[:32],
                description=RootCauseAnalysis.DESCRIPTION_OTHERS,
                description_options=[RootCauseAnalysis.DESCRIPTION_OTHERS],
                status=RootCauseAnalysis.STATUS_DRAFT,
            )
            created_count += 1

    message = f"Created {created_count} draft defect record(s)."
    if blocked_codes:
        message += " Some sub-segments already have ongoing records."
    if not_found_codes:
        message += " Some sub-segments were not found."
    return JsonResponse(
        {
            "ok": True,
            "created_count": created_count,
            "blocked_codes": blocked_codes,
            "not_found_codes": not_found_codes,
            "message": message,
        }
    )


def _filtered_segments_for_road_motorability(request):
    qs = Segment.objects.select_related("route", "start_point", "end_point").all()
    current_view = request.GET.get("view") or "map"
    selected_road = request.GET.get("road") or ""
    selected_route = request.GET.get("route") or ""
    selected_state = (request.GET.get("state") or "").strip()
    selected_speed = (request.GET.get("speed") or "").strip().lower()

    if selected_road:
        selected_route = ""
        selected_state = ""
        selected_speed = ""
        qs = qs.filter(route__road__road=selected_road)
    elif selected_route:
        selected_road = ""
        selected_state = ""
        selected_speed = ""
        qs = qs.filter(route__route=selected_route)
    elif selected_state:
        selected_road = ""
        selected_route = ""
        selected_speed = ""
        qs = qs.filter(state__iexact=selected_state)
    elif selected_speed in STATUS_BUCKETS:
        selected_road = ""
        selected_route = ""
        selected_state = ""
        qs = qs.filter(status__in=STATUS_BUCKETS[selected_speed]["codes"])

    return {
        "qs": qs,
        "current_view": current_view,
        "selected_road": selected_road,
        "selected_route": selected_route,
        "selected_state": selected_state,
        "selected_speed": selected_speed,
    }


def _apply_motorability_filters(qs, selected_road="", selected_route="", selected_state="", selected_speed=""):
    selected_road = selected_road or ""
    selected_route = selected_route or ""
    selected_state = (selected_state or "").strip()
    selected_speed = (selected_speed or "").strip().lower()

    if selected_road:
        selected_route = ""
        selected_state = ""
        selected_speed = ""
        qs = qs.filter(route__road__road=selected_road)
    elif selected_route:
        selected_road = ""
        selected_state = ""
        selected_speed = ""
        qs = qs.filter(route__route=selected_route)
    elif selected_state:
        selected_road = ""
        selected_route = ""
        selected_speed = ""
        qs = qs.filter(state__iexact=selected_state)
    elif selected_speed in STATUS_BUCKETS:
        selected_road = ""
        selected_route = ""
        selected_state = ""
        qs = qs.filter(status__in=STATUS_BUCKETS[selected_speed]["codes"])

    return qs, selected_road, selected_route, selected_state, selected_speed


def _build_road_motorability_context(request):
    filtered = _filtered_segments_for_road_motorability(request)
    qs = filtered["qs"]
    roads = Road.objects.only("road").order_by("road")
    routes = Route.objects.only("route").order_by("route")
    states = State.objects.only("state").order_by("state").values_list("state", flat=True)
    speed_options = [
        ("good", "Good"),
        ("tolerable", "Tolerable"),
        ("intolerable", "Intolerable"),
        ("failed", "Failed"),
        ("no_response", "No response"),
    ]

    paginator = Paginator(qs.order_by("route__route", "index", "code"), 50)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    filters = {}
    if filtered["current_view"]:
        filters["view"] = filtered["current_view"]
    if filtered["selected_road"]:
        filters["road"] = filtered["selected_road"]
    if filtered["selected_route"]:
        filters["route"] = filtered["selected_route"]
    if filtered["selected_state"]:
        filters["state"] = filtered["selected_state"]
    if filtered["selected_speed"]:
        filters["speed"] = filtered["selected_speed"]

    metrics = _overview_metrics(qs)
    all_rows = list(qs.order_by("route__route", "code")[:12])
    focus_segment = all_rows[0] if all_rows else None
    unique_route_count = qs.values("route_id").distinct().count()
    has_rows = qs.exists()

    def _sum_length_or_dash(base_qs):
        total = base_qs.aggregate(total_length=Sum("distance")).get("total_length")
        return _format_km_total(total) if total is not None else "-"

    visible_status_keys = ["good", "tolerable", "intolerable", "failed"]
    kpi_raw_lengths = {
        key: (qs.filter(status__in=STATUS_BUCKETS[key]["codes"]).aggregate(total_length=Sum("distance")).get("total_length") or Decimal("0.00"))
        for key in visible_status_keys
    }
    kpi_denominator = sum(kpi_raw_lengths.values(), Decimal("0.00"))
    kpi_lengths = {}
    for key in visible_status_keys:
        length_value = kpi_raw_lengths[key]
        percent_value = Decimal("0.00")
        if kpi_denominator > 0:
            percent_value = (Decimal(length_value) / kpi_denominator) * Decimal("100")
        kpi_lengths[key] = f"{_format_km_total(length_value)} / {percent_value.quantize(Decimal('0.01')):.2f}%"

    selected_total_length = _sum_length_or_dash(qs) if has_rows else "-"
    good_total_length = (
        _sum_length_or_dash(qs.filter(status__in=STATUS_BUCKETS["good"]["codes"])) if has_rows else "-"
    )
    tolerable_total_length = (
        _sum_length_or_dash(qs.filter(status__in=STATUS_BUCKETS["tolerable"]["codes"])) if has_rows else "-"
    )
    intolerable_total_length = (
        _sum_length_or_dash(qs.filter(status__in=STATUS_BUCKETS["intolerable"]["codes"])) if has_rows else "-"
    )
    failed_total_length = (
        _sum_length_or_dash(qs.filter(status__in=STATUS_BUCKETS["failed"]["codes"])) if has_rows else "-"
    )
    active_defects_qs = (
        Defect.objects.select_related("subsegment__segment")
        .filter(subsegment__segment__in=qs)
        .exclude(workflow_status=Defect.WORKFLOW_REPAIR_COMPLETE)
        .order_by("subsegment__segment__index", "subsegment__segment__code", "-modified", "-id")
    )
    investigation_rows = []
    seen_segment_ids = set()
    for defect in active_defects_qs:
        segment = defect.subsegment.segment if defect.subsegment_id else None
        if not segment or segment.id in seen_segment_ids:
            continue
        seen_segment_ids.add(segment.id)
        investigation_rows.append(
            {
                "subsegment_code": defect.subsegment.code if defect.subsegment_id and defect.subsegment else "-",
                "status_label": defect.get_workflow_status_display(),
            }
        )

    return {
        "active_page": "road_motorability",
        "segments": page_obj.object_list,
        "page_obj": page_obj,
        "roads": roads,
        "routes": routes,
        "states": list(states),
        "speed_options": speed_options,
        "selected_road": filtered["selected_road"],
        "selected_route": filtered["selected_route"],
        "selected_state": filtered["selected_state"],
        "selected_speed": filtered["selected_speed"],
        "current_view": filtered["current_view"],
        "filters_qs": urlencode(filters),
        "number_routes": unique_route_count,
        "segment_length_total": "----",
        "focus_segment": focus_segment,
        "report_rows": all_rows[:3],
        "investigation_rows": investigation_rows,
        "summary_total_length_selected": selected_total_length,
        "summary_number_routes": unique_route_count if has_rows else "-",
        "summary_number_segments": metrics["total_segments"] if has_rows else "-",
        "summary_total_length_good": good_total_length,
        "summary_total_length_tolerable": tolerable_total_length,
        "summary_total_length_intolerable": intolerable_total_length,
        "summary_total_length_failed": failed_total_length,
        "kpi_good_display": kpi_lengths["good"],
        "kpi_tolerable_display": kpi_lengths["tolerable"],
        "kpi_intolerable_display": kpi_lengths["intolerable"],
        "kpi_failed_display": kpi_lengths["failed"],
        **metrics,
        "counts": kpi_lengths,
    }


def road_motorability(request):
    context = _build_road_motorability_context(request)
    return render(request, "website/road_motorability.html", context)


def road_motorability_queue_refresh(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "POST method required."}, status=405)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        payload = {}

    selected_road = (payload.get("road") or "").strip()
    selected_route = (payload.get("route") or "").strip()
    selected_state = (payload.get("state") or "").strip()
    selected_speed = (payload.get("speed") or "").strip().lower()
    refreshed_codes = payload.get("refreshed_codes") or []
    if not isinstance(refreshed_codes, list):
        refreshed_codes = []
    refreshed_set = {str(code).strip().upper() for code in refreshed_codes if str(code).strip()}

    base_qs = Segment.objects.all()
    filtered_qs, _, _, _, _ = _apply_motorability_filters(
        base_qs,
        selected_road=selected_road,
        selected_route=selected_route,
        selected_state=selected_state,
        selected_speed=selected_speed,
    )
    all_codes = [str(code).strip().upper() for code in filtered_qs.values_list("code", flat=True) if str(code).strip()]
    already_refreshed_codes = [code for code in all_codes if code in refreshed_set]
    to_queue = [code for code in all_codes if code not in refreshed_set]

    if not to_queue:
        return JsonResponse(
            {
                "ok": True,
                "queued": False,
                "task_id": "",
                "queued_count": 0,
                "total_filtered": len(all_codes),
                "already_refreshed": len(already_refreshed_codes),
                "already_refreshed_codes": already_refreshed_codes,
                "queued_codes": [],
                "message": "No eligible segments to refresh for current filters.",
            }
        )

    try:
        async_result = refresh_segments_task.delay(codes=to_queue)
    except Exception as exc:
        return JsonResponse(
            {
                "ok": False,
                "message": f"Could not queue motorability refresh. ({exc})",
            },
            status=500,
        )
    return JsonResponse(
        {
            "ok": True,
            "queued": True,
            "task_id": str(async_result.id),
            "queued_count": len(to_queue),
            "total_filtered": len(all_codes),
            "already_refreshed": len(already_refreshed_codes),
            "already_refreshed_codes": already_refreshed_codes,
            "queued_codes": to_queue,
            "message": f"Queued {len(to_queue)} segments for refresh.",
        }
    )

def _segments_geojson_response(segments):
    features = []

    for seg in segments:
        try:
            start_lat = float(seg.start_lat)
            start_lon = float(seg.start_lon)
            end_lat = float(seg.end_lat)
            end_lon = float(seg.end_lon)
        except (TypeError, ValueError, InvalidOperation):
            continue

        if start_lat == end_lat and start_lon == end_lon:
            geometry = {"type": "Point", "coordinates": [start_lon, start_lat]}
        else:
            geometry = {
                "type": "LineString",
                "coordinates": [[start_lon, start_lat], [end_lon, end_lat]],
            }

        features.append(
            {
                "type": "Feature",
                "geometry": geometry,
                "properties": {
                    "code": seg.code,
                    "route": getattr(seg.route, "route", ""),
                    "state": seg.state or "",
                    "distance": float(seg.distance or 0),
                    "avg_speed": float(seg.avg_speed or 0),
                    "status": seg.status or "666699",
                },
            }
        )

    return JsonResponse({"type": "FeatureCollection", "features": features})


def segments_map_data(request):
    filtered = _filtered_segments_for_road_motorability(request)
    segments = filtered["qs"].order_by("route__route", "index", "code")
    return _segments_geojson_response(segments)


def library_landing(request, active_section="road_inventory"):
    qs = Segment.objects.select_related("route", "start_point", "end_point").all()
    selected_road = (request.GET.get("road") or "").strip()
    selected_route = (request.GET.get("route") or "").strip()
    selected_state = (request.GET.get("state") or "").strip()
    segment_code_query = (request.GET.get("segment_code") or "").strip()

    if selected_road:
        qs = qs.filter(route__road__road=selected_road)
    if selected_route:
        qs = qs.filter(route__route=selected_route)
    if selected_state:
        qs = qs.filter(state__iexact=selected_state)
    if segment_code_query:
        qs = qs.filter(code__icontains=segment_code_query)

    qs = qs.order_by("route__route", "index", "code")
    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    qd = request.GET.copy()
    qd.pop("page", None)
    filters_qs = qd.urlencode()
    documentation = _load_library_documentation((request.GET.get("doc") or "").strip())

    return render(
        request,
        "website/library_landing.html",
        {
            "active_page": "library",
            "active_library_section": active_section,
            "segments": page_obj.object_list,
            "page_obj": page_obj,
            "filters_qs": filters_qs,
            "roads": Road.objects.only("road").order_by("road"),
            "routes": Route.objects.only("route").order_by("route"),
            "states": State.objects.only("state").order_by("state"),
            "selected_road": selected_road,
            "selected_route": selected_route,
            "selected_state": selected_state,
            "segment_code_query": segment_code_query,
            **documentation,
        },
    )


def _library_segment_editor_payload(segment):
    return {
        "route": segment.route.route if segment.route_id and segment.route else "",
        "segment_code": segment.code or "",
        "name": segment.name or "",
        "state": segment.state or "",
        "start_name": segment.start_point.name if segment.start_point_id and segment.start_point and segment.start_point.name else "",
        "northings": str(segment.start_lat) if segment.start_lat is not None else "",
        "eastings": str(segment.start_lon) if segment.start_lon is not None else "",
        "end_name": segment.end_point.name if segment.end_point_id and segment.end_point and segment.end_point.name else "",
        "northings_2": str(segment.end_lat) if segment.end_lat is not None else "",
        "eastings_2": str(segment.end_lon) if segment.end_lon is not None else "",
    }


def _resolve_library_address(name, lat, lon):
    clean_name = (name or "").strip()
    if not clean_name:
        return None
    address = Address.objects.filter(name__iexact=clean_name).order_by("id").first()
    if address is None:
        return Address.objects.create(
            address=clean_name,
            name=clean_name,
            lat=lat if lat is not None else Decimal("0"),
            lng=lon if lon is not None else Decimal("0"),
        )
    fields_to_update = []
    if address.name != clean_name:
        address.name = clean_name
        fields_to_update.append("name")
    if lat is not None and address.lat != lat:
        address.lat = lat
        fields_to_update.append("lat")
    if lon is not None and address.lng != lon:
        address.lng = lon
        fields_to_update.append("lng")
    if fields_to_update:
        address.save(update_fields=fields_to_update)
    return address


def library_segment_editor(request, segment_code):
    segment = (
        Segment.objects.select_related("route", "start_point", "end_point")
        .filter(code__iexact=(segment_code or "").strip())
        .first()
    )
    if segment is None:
        return JsonResponse({"ok": False, "message": "Segment not found."}, status=404)

    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "POST method required."}, status=405)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except (TypeError, ValueError, json.JSONDecodeError):
        return JsonResponse({"ok": False, "message": "Invalid request payload."}, status=400)

    route_code = str(payload.get("route") or "").strip().upper()
    new_segment_code = str(payload.get("segment_code") or "").strip().upper()
    name = str(payload.get("name") or "").strip()
    state = str(payload.get("state") or "").strip()
    start_name = str(payload.get("start_name") or "").strip()
    end_name = str(payload.get("end_name") or "").strip()
    start_lat = _to_decimal_or_none(payload.get("northings"))
    start_lon = _to_decimal_or_none(payload.get("eastings"))
    end_lat = _to_decimal_or_none(payload.get("northings_2"))
    end_lon = _to_decimal_or_none(payload.get("eastings_2"))
    if not route_code:
        return JsonResponse({"ok": False, "message": "Route is required."}, status=400)
    if not new_segment_code:
        return JsonResponse({"ok": False, "message": "Segment code is required."}, status=400)

    duplicate_exists = (
        Segment.objects.exclude(pk=segment.pk)
        .filter(code__iexact=new_segment_code)
        .exists()
    )
    if duplicate_exists:
        return JsonResponse({"ok": False, "message": f'Segment code "{new_segment_code}" already exists.'}, status=400)

    with transaction.atomic():
        road_code = _road_code_from_route(route_code)
        road_obj, _ = Road.objects.get_or_create(road=road_code)
        route_obj, _ = Route.objects.get_or_create(route=route_code, defaults={"road": road_obj})
        if route_obj.road_id != road_obj.id:
            route_obj.road = road_obj
            route_obj.save(update_fields=["road"])

        segment.route = route_obj
        segment.code = new_segment_code
        segment.name = name
        segment.state = state
        segment.start_lat = start_lat if start_lat is not None else Decimal("0")
        segment.start_lon = start_lon if start_lon is not None else Decimal("0")
        segment.end_lat = end_lat if end_lat is not None else Decimal("0")
        segment.end_lon = end_lon if end_lon is not None else Decimal("0")

        if start_name:
            segment.start_point = _resolve_library_address(start_name, segment.start_lat, segment.start_lon)
        if end_name:
            segment.end_point = _resolve_library_address(end_name, segment.end_lat, segment.end_lon)

        segment.save()

    segment.refresh_from_db()
    return JsonResponse({"ok": True, "segment": _library_segment_editor_payload(segment)})


def library_segments_bulk_delete(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "POST method required."}, status=405)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except (TypeError, ValueError, json.JSONDecodeError):
        return JsonResponse({"ok": False, "message": "Invalid request payload."}, status=400)

    raw_codes = payload.get("segment_codes") or []
    segment_codes = []
    for code in raw_codes:
        cleaned = str(code or "").strip().upper()
        if cleaned and cleaned not in segment_codes:
            segment_codes.append(cleaned)

    if not segment_codes:
        return JsonResponse({"ok": False, "message": "No segment codes were provided."}, status=400)

    deleted_codes = []
    failures = []

    for code in segment_codes:
        segment = Segment.objects.filter(code__iexact=code).first()
        if segment is None:
            failures.append({"segment_code": code, "reason": "Segment was not found."})
            continue
        try:
            with transaction.atomic():
                segment.delete()
            deleted_codes.append(code)
        except ProtectedError as exc:
            protected_labels = []
            for obj in list(exc.protected_objects)[:5]:
                protected_labels.append(f"{obj.__class__.__name__} #{getattr(obj, 'pk', '?')}")
            detail = ", ".join(protected_labels) if protected_labels else "Protected related records still exist."
            failures.append(
                {
                    "segment_code": code,
                    "reason": f"Deletion blocked by related records: {detail}",
                }
            )
        except Exception as exc:
            failures.append({"segment_code": code, "reason": str(exc) or "Unknown deletion error."})

    return JsonResponse(
        {
            "ok": True,
            "deleted_codes": deleted_codes,
            "failures": failures,
        }
    )


def library_reports(request):
    def _status_label_and_class(status_code):
        if status_code == "FF5050":
            return "Failed", "is-intolerable"
        if status_code == "FF9966":
            return "Intolerable", "is-intolerable"
        if status_code in {"00CC00", "05700B"}:
            return "Tolerable", "is-tolerable"
        return "No response", "is-neutral"

    report_rows = []
    selected_state = (request.GET.get("state") or "").strip()

    rca_qs = (
        RootCauseAnalysis.objects.select_related("subsegment", "subsegment__segment")
        .prefetch_related("library_files")
        .annotate(
            library_attachments_count=Count(
                "library_files",
                filter=Q(library_files__entry_type=Library.TYPE_ROOT_CAUSE_ANALYSIS),
            )
        )
        .order_by("-id")
    )
    if selected_state:
        rca_qs = rca_qs.filter(subsegment__segment__state__iexact=selected_state)
    for idx, report in enumerate(rca_qs):
        segment = getattr(report.subsegment, "segment", None)
        status_code = getattr(segment, "status", "666699")
        status_label, status_class = _status_label_and_class(status_code)
        linked_file = (
            report.library_files.filter(entry_type=Library.TYPE_ROOT_CAUSE_ANALYSIS)
            .order_by("-created", "-id")
            .first()
        )
        report_rows.append(
            {
                "file_name": report.subsegment.code if report.subsegment else "-",
                "report_type": "Root cause report",
                "road_condition": status_label,
                "road_condition_class": status_class,
                "last_updated": f"{timesince(report.updated_at).split(',')[0]} ago",
                "uploaded_by": "Engineer Ridwan Bankole",
                "attachments_count": (report.library_attachments_count or 0),
                "attachment_url": linked_file.file.url if linked_file else "",
                "attachment_name": linked_file.name if linked_file else "",
                "details_url": f"{reverse('engineering_admin')}?mode=view&analysis={report.pk}",
            }
        )

    physical_qs = (
        PhysicalInspection.objects.select_related("subsegment", "subsegment__segment")
        .prefetch_related("library_files")
        .annotate(
            library_attachments_count=Count(
                "library_files",
                filter=Q(library_files__entry_type=Library.TYPE_PHYSICAL_INSPECTION),
            )
        )
        .order_by("-updated_at", "-id")
    )
    if selected_state:
        physical_qs = physical_qs.filter(subsegment__segment__state__iexact=selected_state)
    for idx, report in enumerate(physical_qs):
        segment = getattr(report.subsegment, "segment", None)
        status_code = getattr(segment, "status", "666699")
        status_label, status_class = _status_label_and_class(status_code)
        linked_file = (
            report.library_files.filter(entry_type=Library.TYPE_PHYSICAL_INSPECTION)
            .order_by("-created", "-id")
            .first()
        )
        report_rows.append(
            {
                "file_name": report.subsegment.code if report.subsegment else "-",
                "report_type": "Physical inspection report",
                "road_condition": status_label,
                "road_condition_class": status_class,
                "last_updated": f"{timesince(report.updated_at).split(',')[0]} ago",
                "uploaded_by": "Engineer Ridwan Bankole",
                "attachments_count": (report.library_attachments_count or 0),
                "attachment_url": linked_file.file.url if linked_file else "",
                "attachment_name": linked_file.name if linked_file else "",
                "details_url": f"{reverse('physical_inspection')}?mode=view&inspection={report.pk}",
            }
        )

    report_rows.sort(key=lambda row: row["file_name"])
    paginator = Paginator(report_rows, 50)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    qd = request.GET.copy()
    qd.pop("page", None)
    filters_qs = qd.urlencode()
    documentation = _load_library_documentation((request.GET.get("doc") or "").strip())

    return render(
        request,
        "website/library_reports.html",
        {
            "active_page": "library",
            "active_library_section": "reports",
            "report_rows": page_obj.object_list,
            "page_obj": page_obj,
            "states": State.objects.only("state").order_by("state"),
            "selected_state": selected_state,
            "filters_qs": filters_qs,
            **documentation,
        },
    )


def _render_library_guide(request, *, entry_type, active_section, search_label, url_name):
    query = (request.GET.get("q") or "").strip()
    guides_qs = Library.objects.filter(entry_type=entry_type).order_by("-created", "-id")
    if query:
        guides_qs = guides_qs.filter(name__icontains=query)
    documentation = _load_library_documentation((request.GET.get("doc") or "").strip())

    return render(
        request,
        "website/library_technical_guide.html",
        {
            "active_page": "library",
            "active_library_section": active_section,
            "guide_documents": guides_qs,
            "guide_documents_count": guides_qs.count(),
            "search_query": query,
            "guide_search_label": search_label,
            "guide_url_name": url_name,
            **documentation,
        },
    )


def library_technical_guide(request):
    return _render_library_guide(
        request,
        entry_type=Library.TYPE_TECHNICAL_GUIDE,
        active_section="technical_guide",
        search_label="technical guide",
        url_name="library_technical_guide",
    )


def library_user_guide(request):
    return _render_library_guide(
        request,
        entry_type=Library.TYPE_USER_GUIDE,
        active_section="user_guide",
        search_label="user guide",
        url_name="library_user_guide",
    )


def engineering_admin_root_cause(request):
    origin_key = request.GET.get("origin") or request.POST.get("origin") or ""
    back_link_text, back_link_href = _resolve_origin_back_link(
        origin_key, "Root Cause Analysis", reverse("engineering_admin")
    )
    mode_value = request.GET.get("mode")
    if mode_value == "form":
        library_mode = "form"
    elif mode_value == "view":
        library_mode = "view"
    else:
        library_mode = "summary"
    analysis_id = request.GET.get("analysis") or request.POST.get("analysis_id")
    defect_id = request.GET.get("defect") or request.POST.get("defect_id")
    subsegment_id = request.GET.get("subsegment") or request.POST.get("subsegment_id")
    subsegments_qs = SubSegment.objects.select_related("segment").order_by("segment_id", "position")
    selected_defect = None
    if defect_id:
        selected_defect = Defect.objects.select_related("subsegment", "engineer").filter(pk=defect_id).first()
    selected_subsegment = None
    if selected_defect:
        selected_subsegment = selected_defect.subsegment
    elif subsegment_id:
        selected_subsegment = subsegments_qs.filter(pk=subsegment_id).first()
    if selected_subsegment is None:
        selected_subsegment = subsegments_qs.first()
    defect_type_qs = DefectType.objects.filter(is_active=True).order_by("label")
    description_options = list(defect_type_qs.values_list("code", "label"))
    if not description_options:
        description_options = list(RootCauseAnalysis.DESCRIPTION_CHOICES)
    existing_analysis = None
    if selected_defect:
        existing_analysis = (
            RootCauseAnalysis.objects.prefetch_related(
                "defect_types", "root_cause_details", "library_files"
            )
            .filter(defect=selected_defect)
            .order_by("-updated_at", "-id")
            .first()
        )
    elif analysis_id:
        existing_analysis = (
            RootCauseAnalysis.objects.prefetch_related(
                "defect_types", "root_cause_details", "library_files"
            )
            .filter(pk=analysis_id)
            .first()
        )
        if existing_analysis and selected_defect is None:
            selected_defect = existing_analysis.defect
        if existing_analysis and selected_subsegment is None:
            selected_subsegment = existing_analysis.subsegment

    rca_values = {
        "description": request.POST.get("description", ""),
        "subgrade_properties": request.POST.get("subgrade_properties", ""),
        "vegetation": request.POST.get("vegetation", ""),
        "topography": request.POST.get("topography", ""),
        "drainage_characteristics": request.POST.get("drainage_characteristics", ""),
        "temp_humidity": request.POST.get("temp_humidity", ""),
        "location": request.POST.get("location", ""),
    }
    selected_descriptions = request.POST.getlist("description_options")
    rca_error = ""
    rca_success = request.GET.get("saved") == "1"
    if request.method != "POST" and library_mode == "form" and existing_analysis:
        selected_subsegment = existing_analysis.subsegment
        selected_descriptions = list(existing_analysis.defect_types.values_list("code", flat=True))
        if not selected_descriptions:
            selected_descriptions = existing_analysis.description_options or []
        if not selected_descriptions and existing_analysis.description:
            selected_descriptions = [existing_analysis.description]
        rca_values["location"] = existing_analysis.location
        feature_key_map = {
            RootCauseDetail.FEATURE_SUBGRADE_PROPERTIES: "subgrade_properties",
            RootCauseDetail.FEATURE_VEGETATION: "vegetation",
            RootCauseDetail.FEATURE_TOPOGRAPHY: "topography",
            RootCauseDetail.FEATURE_DRAINAGE_CHARACTERISTICS: "drainage_characteristics",
            RootCauseDetail.FEATURE_TEMPERATURE_HUMIDITY: "temp_humidity",
        }
        for detail in existing_analysis.root_cause_details.all():
            feature_key = feature_key_map.get(detail.natural_feature)
            if feature_key:
                rca_values[feature_key] = detail.characteristic

    if request.method == "POST" and library_mode == "form":
        submit_action = request.POST.get("submit_action")
        if submit_action == "complete":
            status_value = RootCauseAnalysis.STATUS_COMPLETE
        else:
            status_value = RootCauseAnalysis.STATUS_DRAFT
        supporting_file = request.FILES.get("supporting_file")

        if selected_subsegment is None:
            rca_error = "No subsegment is available. Add subsegment data before creating root cause analysis."
        elif not selected_descriptions:
            rca_error = "Select at least one defect description."
        else:
            valid_description_values = {value for value, _ in description_options}
            invalid_descriptions = [value for value in selected_descriptions if value not in valid_description_values]
            if invalid_descriptions:
                rca_error = "One or more selected defect descriptions are invalid."
                selected_primary_description = ""
            else:
                selected_primary_description = selected_descriptions[0]

            feature_pairs = [
                (RootCauseDetail.FEATURE_SUBGRADE_PROPERTIES, rca_values["subgrade_properties"]),
                (RootCauseDetail.FEATURE_VEGETATION, rca_values["vegetation"]),
                (RootCauseDetail.FEATURE_TOPOGRAPHY, rca_values["topography"]),
                (RootCauseDetail.FEATURE_DRAINAGE_CHARACTERISTICS, rca_values["drainage_characteristics"]),
                (RootCauseDetail.FEATURE_TEMPERATURE_HUMIDITY, rca_values["temp_humidity"]),
            ]
            selected_feature_pairs = [(feature, value) for feature, value in feature_pairs if value]
            valid_by_feature = {
                feature: {value for value, _ in choices}
                for feature, choices in RootCauseDetail.CHARACTERISTIC_CHOICES_BY_FEATURE.items()
            }
            invalid_choice = any(
                value not in valid_by_feature.get(feature, set())
                for feature, value in selected_feature_pairs
            )
            if invalid_choice and not rca_error:
                rca_error = "One or more selected characteristic values are invalid."
            if not rca_error:
                if selected_defect and selected_defect.subsegment_id == selected_subsegment.id:
                    defect_for_analysis = selected_defect
                else:
                    try:
                        defect_for_analysis, _ = _get_or_create_active_defect(selected_subsegment)
                    except ValueError as exc:
                        rca_error = str(exc)
                if not rca_error:
                    with transaction.atomic():
                        location_value = (
                            (rca_values["location"] or "").strip()
                            or (selected_subsegment.code or "")[:32]
                            or "Unknown"
                        )
                        if existing_analysis:
                            analysis = existing_analysis
                            analysis.subsegment = selected_subsegment
                            analysis.defect = defect_for_analysis
                            analysis.location = location_value[:32]
                            analysis.description = selected_primary_description
                            analysis.description_options = selected_descriptions
                            analysis.status = status_value
                            analysis.save()
                        else:
                            analysis = RootCauseAnalysis.objects.create(
                                subsegment=selected_subsegment,
                                defect=defect_for_analysis,
                                location=location_value[:32],
                                description=selected_primary_description,
                                description_options=selected_descriptions,
                                status=status_value,
                            )
                        if supporting_file:
                            Library.objects.create(
                                entry_type=Library.TYPE_ROOT_CAUSE_ANALYSIS,
                                file_type=_library_file_type_from_name(supporting_file.name),
                                name=os.path.basename(supporting_file.name),
                                file=supporting_file,
                                defect=defect_for_analysis,
                                root_cause_analysis=analysis,
                            )
                        selected_defect_types = list(
                            defect_type_qs.filter(code__in=selected_descriptions)
                        )
                        analysis.defect_types.set(selected_defect_types)
                        selected_features = set()
                        for feature, value in selected_feature_pairs:
                            RootCauseDetail.objects.update_or_create(
                                root_cause_analysis=analysis,
                                natural_feature=feature,
                                defaults={
                                    "characteristic": value,
                                    "root_cause_analysis_text": "",
                                },
                            )
                            selected_features.add(feature)
                        if selected_features:
                            analysis.root_cause_details.exclude(
                                natural_feature__in=selected_features
                            ).delete()
                        else:
                            analysis.root_cause_details.all().delete()
                        _update_defect_status_from_rca(defect_for_analysis, status_value)
                        _touch_defect_modified(defect_for_analysis)

                    if status_value == RootCauseAnalysis.STATUS_COMPLETE:
                        _ensure_physical_draft_for_defect(defect_for_analysis)
                    return redirect(reverse("engineering_admin"))

    root_cause_analyses = (
        RootCauseAnalysis.objects.select_related("subsegment", "subsegment__segment", "defect")
        .prefetch_related("defect_types", "root_cause_details", "library_files")
        .order_by("-updated_at", "-id")
    )
    rca_table_statuses = [Defect.WORKFLOW_DRAFT, Defect.WORKFLOW_RCA]
    defects_qs = (
        Defect.objects.select_related("subsegment", "subsegment__segment", "engineer")
        .filter(workflow_status__in=rca_table_statuses)
        .order_by("-modified", "-id")
    )
    root_cause_total = defects_qs.count()
    root_cause_draft_count = defects_qs.filter(workflow_status=Defect.WORKFLOW_DRAFT).count()
    root_cause_complete_count = defects_qs.filter(workflow_status=Defect.WORKFLOW_RCA).count()
    description_label_map = dict(description_options)
    latest_analysis_by_defect = {}
    for report in root_cause_analyses:
        if report.defect_id and report.defect_id not in latest_analysis_by_defect:
            latest_analysis_by_defect[report.defect_id] = report

    draft_rows = []
    complete_rows = []
    condition_map = {
        Defect.CONDITION_TOLERABLE: ("Tolerable", "tolerable"),
        Defect.CONDITION_INTOLERABLE: ("Intolerable", "intolerable"),
        Defect.CONDITION_FAILED: ("Intolerable", "intolerable"),
        Defect.CONDITION_BAD: ("No response", "neutral"),
    }
    for defect in defects_qs[:100]:
        report = latest_analysis_by_defect.get(defect.id)
        if report:
            selected_values = list(report.defect_types.values_list("code", flat=True))
            if not selected_values:
                selected_values = report.description_options or []
            selected_labels = [description_label_map.get(value) for value in selected_values if value in description_label_map]
            defect_text = ", ".join(label.lower() for label in selected_labels if label) or report.get_description_display()
            date_text = f"{timesince(report.updated_at).split(',')[0]} ago"
        else:
            defect_text = "not specified"
            date_text = f"{timesince(defect.modified).split(',')[0]} ago"
        condition_label, condition_class = condition_map.get(defect.condition, ("No response", "neutral"))
        engineer_name = "Unassigned"
        if defect.engineer:
            full_name = f"{defect.engineer.first_name} {defect.engineer.last_name}".strip()
            engineer_name = full_name or defect.engineer.get_username()
        row = {
            "defect": defect,
            "report": report,
            "defect_text": defect_text,
            "status_text": defect.get_workflow_status_display(),
            "condition_label": condition_label,
            "condition_class": condition_class,
            "engineer_name": engineer_name,
            "date_text": date_text,
        }
        if defect.workflow_status == Defect.WORKFLOW_DRAFT:
            draft_rows.append(row)
        else:
            complete_rows.append(row)

    if library_mode == "view" and existing_analysis is None:
        if selected_defect:
            existing_analysis = latest_analysis_by_defect.get(selected_defect.id)
        if existing_analysis is None and complete_rows:
            existing_analysis = complete_rows[0].get("report")

    view_defect_text = ""
    view_feature_values = {
        "subgrade_properties": "Not provided",
        "vegetation": "Not provided",
        "topography": "Not provided",
        "drainage_characteristics": "Not provided",
        "temp_humidity": "Not provided",
    }
    supporting_documents = []
    view_segment_status_text = "Unknown"
    view_segment_status_class = "neutral"
    if existing_analysis:
        selected_values = list(existing_analysis.defect_types.values_list("code", flat=True))
        if not selected_values:
            selected_values = existing_analysis.description_options or []
        selected_labels = [description_label_map.get(value) for value in selected_values if value in description_label_map]
        view_defect_text = ", ".join(selected_labels) or existing_analysis.get_description_display()

        feature_key_map = {
            RootCauseDetail.FEATURE_SUBGRADE_PROPERTIES: "subgrade_properties",
            RootCauseDetail.FEATURE_VEGETATION: "vegetation",
            RootCauseDetail.FEATURE_TOPOGRAPHY: "topography",
            RootCauseDetail.FEATURE_DRAINAGE_CHARACTERISTICS: "drainage_characteristics",
            RootCauseDetail.FEATURE_TEMPERATURE_HUMIDITY: "temp_humidity",
        }
        for detail in existing_analysis.root_cause_details.all():
            characteristic_value = detail.get_characteristic_display()
            feature_key = feature_key_map.get(detail.natural_feature)
            if feature_key:
                view_feature_values[feature_key] = characteristic_value

        rca_attachments = list(
            existing_analysis.library_files.filter(
                entry_type=Library.TYPE_ROOT_CAUSE_ANALYSIS
            ).order_by("-created", "-id")
        )
        if rca_attachments:
            supporting_documents = [item.file.name.split("/")[-1] for item in rca_attachments]
        else:
            code = existing_analysis.subsegment.code if existing_analysis.subsegment else "A1LAS2-01"
            supporting_documents = [f"{code} Root Cause.jpg", f"{code} Root Cause.jpg"]

        segment_status_code = getattr(getattr(existing_analysis.subsegment, "segment", None), "status", "")
        if segment_status_code == "FF5050":
            view_segment_status_text = "Failed"
            view_segment_status_class = "intolerable"
        elif segment_status_code == "FF9966":
            view_segment_status_text = "Intolerable"
            view_segment_status_class = "intolerable"
        elif segment_status_code in {"00CC00", "05700B"}:
            view_segment_status_text = "Tolerable"
            view_segment_status_class = "tolerable"
        else:
            view_segment_status_text = "No response"
            view_segment_status_class = "neutral"

    return render(
        request,
        "website/engineering_admin.html",
        {
            "active_page": "engineering_admin",
            "active_library_tab": "approvals" if origin_key == "approvals" else "root_cause",
            "active_content_tab": "root_cause",
            "back_link_text": back_link_text,
            "back_link_href": back_link_href,
            "library_mode": library_mode,
            "rca_error": rca_error,
            "rca_success": rca_success,
            "subsegments": subsegments_qs[:300],
            "selected_subsegment": selected_subsegment,
            "selected_defect": selected_defect,
            "description_options": description_options,
            "selected_descriptions": selected_descriptions,
            "subgrade_options": RootCauseDetail.CHARACTERISTIC_CHOICES_BY_FEATURE[
                RootCauseDetail.FEATURE_SUBGRADE_PROPERTIES
            ],
            "vegetation_options": RootCauseDetail.CHARACTERISTIC_CHOICES_BY_FEATURE[
                RootCauseDetail.FEATURE_VEGETATION
            ],
            "topography_options": RootCauseDetail.CHARACTERISTIC_CHOICES_BY_FEATURE[
                RootCauseDetail.FEATURE_TOPOGRAPHY
            ],
            "drainage_options": RootCauseDetail.CHARACTERISTIC_CHOICES_BY_FEATURE[
                RootCauseDetail.FEATURE_DRAINAGE_CHARACTERISTICS
            ],
            "temperature_humidity_options": RootCauseDetail.CHARACTERISTIC_CHOICES_BY_FEATURE[
                RootCauseDetail.FEATURE_TEMPERATURE_HUMIDITY
            ],
            "rca_values": rca_values,
            "current_analysis": existing_analysis,
            "draft_rows": draft_rows,
            "complete_rows": complete_rows,
            "root_cause_total": root_cause_total,
            "root_cause_draft_count": root_cause_draft_count,
            "root_cause_complete_count": root_cause_complete_count,
            "view_defect_text": view_defect_text,
            "view_feature_values": view_feature_values,
            "supporting_documents": supporting_documents,
            "view_segment_status_text": view_segment_status_text,
            "view_segment_status_class": view_segment_status_class,
        },
    )


def engineering_admin_overview(request):
    return render(
        request,
        "website/engineering_admin.html",
        {
            "active_page": "engineering_admin",
            "active_library_tab": "overview",
        },
    )


def physical_inspection(request):
    active_physical_tab = (
        "physical_2"
        if getattr(getattr(request, "resolver_match", None), "url_name", "") == "physical_inspection_2"
        else "physical"
    )
    origin_key = request.GET.get("origin") or request.POST.get("origin") or ""
    default_physical_route = "physical_inspection_2" if active_physical_tab == "physical_2" else "physical_inspection"
    default_physical_label = "Physical Inspection" if active_physical_tab == "physical" else "Physical 2"
    back_link_text, back_link_href = _resolve_origin_back_link(
        origin_key, default_physical_label, reverse(default_physical_route)
    )
    mode_value = request.GET.get("mode") or request.POST.get("mode")
    if mode_value == "form":
        library_mode = "form"
    elif mode_value == "view":
        library_mode = "view"
    else:
        library_mode = "summary"
    physical_success = request.GET.get("saved") == "1"
    physical_error = ""
    inspection_id = request.GET.get("inspection") or request.POST.get("inspection_id")
    defect_options = list(
        DefectType.objects.filter(is_active=True).order_by("label").values_list("code", "label")
    )
    if not defect_options:
        defect_options = list(RootCauseAnalysis.DESCRIPTION_CHOICES)

    horizontal_alignment_options = [
        value
        for value in PhysicalInspectionCharacteristic.CHARACTERISTICS_BY_OPTION[
            PhysicalInspectionAnalysis.OPTION_HORIZONTAL_ALIGNMENT
        ]
    ]
    vertical_alignment_options = [
        value
        for value in PhysicalInspectionCharacteristic.CHARACTERISTICS_BY_OPTION[
            PhysicalInspectionAnalysis.OPTION_VERTICAL_ALIGNMENT
        ]
    ]
    bridge_options = [
        value
        for value in PhysicalInspectionCharacteristic.CHARACTERISTICS_BY_OPTION[
            PhysicalInspectionAnalysis.OPTION_BRIDGES
        ]
    ]

    selected_subsegment = None
    selected_subsegment_id = request.GET.get("subsegment") or request.POST.get("subsegment_id")
    if selected_subsegment_id:
        selected_subsegment = SubSegment.objects.filter(pk=selected_subsegment_id).first()

    selected_defect = None
    selected_defect_id = request.GET.get("defect") or request.POST.get("defect_id")
    if selected_defect_id:
        selected_defect = (
            Defect.objects.select_related("subsegment", "subsegment__segment", "engineer")
            .filter(pk=selected_defect_id)
            .first()
        )
        if selected_defect and selected_subsegment is None:
            selected_subsegment = selected_defect.subsegment

    existing_inspection = None
    if inspection_id:
        existing_inspection = (
            PhysicalInspection.objects.select_related("subsegment", "subsegment__segment")
            .prefetch_related("defect_types", "analysis_rows__characteristics")
            .filter(pk=inspection_id)
            .first()
        )
    elif selected_defect:
        existing_inspection = (
            PhysicalInspection.objects.select_related("subsegment", "subsegment__segment")
            .prefetch_related("defect_types", "analysis_rows__characteristics")
            .filter(defect=selected_defect)
            .order_by("-updated_at", "-id")
            .first()
        )

    physical_values = {
        "segment_id": request.POST.get("segment_id", "").strip(),
        "horizontal_alignment": request.POST.get("horizontal_alignment", "").strip(),
        "vertical_alignment": request.POST.get("vertical_alignment", "").strip(),
        "bridges": request.POST.get("bridges", "").strip(),
        "selected_defects": request.POST.getlist("defect_types"),
    }
    if request.method != "POST" and library_mode == "form" and existing_inspection:
        physical_values["segment_id"] = existing_inspection.subsegment.code
        physical_values["selected_defects"] = list(
            existing_inspection.defect_types.values_list("code", flat=True)
        )
        option_rows = {row.option: row for row in existing_inspection.analysis_rows.all()}
        horizontal_row = option_rows.get(PhysicalInspectionAnalysis.OPTION_HORIZONTAL_ALIGNMENT)
        if horizontal_row:
            first = horizontal_row.characteristics.first()
            if first:
                physical_values["horizontal_alignment"] = first.value or first.characteristic
        vertical_row = option_rows.get(PhysicalInspectionAnalysis.OPTION_VERTICAL_ALIGNMENT)
        if vertical_row:
            first = vertical_row.characteristics.first()
            if first:
                physical_values["vertical_alignment"] = first.value or first.characteristic
        bridges_row = option_rows.get(PhysicalInspectionAnalysis.OPTION_BRIDGES)
        if bridges_row:
            first = bridges_row.characteristics.first()
            if first:
                physical_values["bridges"] = first.value or first.characteristic
    elif request.method != "POST" and library_mode == "form" and selected_subsegment:
        physical_values["segment_id"] = selected_subsegment.code

    if request.method == "POST" and library_mode == "form":
        submit_action = request.POST.get("submit_action")
        status_value = (
            PhysicalInspection.STATUS_COMPLETE
            if submit_action == "complete"
            else PhysicalInspection.STATUS_DRAFT
        )
        subsegment_code = physical_values["segment_id"]
        selected_defects = physical_values["selected_defects"]
        valid_defect_codes = {value for value, _ in defect_options}
        selected_defects = [code for code in selected_defects if code in valid_defect_codes]

        if not subsegment_code:
            physical_error = "Segment id is required."
        else:
            selected_subsegment = SubSegment.objects.filter(code__iexact=subsegment_code).first()
            if selected_subsegment is None:
                physical_error = "Segment id was not found."
            elif not selected_defects:
                physical_error = "Select at least one defect description."
            else:
                valid_horizontal = set(horizontal_alignment_options)
                valid_vertical = set(vertical_alignment_options)
                valid_bridges = set(bridge_options)
                horizontal_value = physical_values["horizontal_alignment"]
                vertical_value = physical_values["vertical_alignment"]
                bridges_value = physical_values["bridges"]
                if horizontal_value and horizontal_value not in valid_horizontal:
                    physical_error = "Invalid horizontal alignment value."
                elif vertical_value and vertical_value not in valid_vertical:
                    physical_error = "Invalid vertical alignment value."
                elif bridges_value and bridges_value not in valid_bridges:
                    physical_error = "Invalid bridges value."

        if not physical_error:
            if selected_defect and selected_defect.subsegment_id == selected_subsegment.id:
                defect_for_inspection = selected_defect
            else:
                try:
                    defect_for_inspection, _ = _get_or_create_active_defect(selected_subsegment)
                except ValueError as exc:
                    physical_error = str(exc)
            if not physical_error:
                with transaction.atomic():
                    if existing_inspection:
                        inspection = existing_inspection
                        inspection.subsegment = selected_subsegment
                        inspection.defect = defect_for_inspection
                        inspection.status = status_value
                        inspection.save()
                        inspection.analysis_rows.all().delete()
                    else:
                        inspection = PhysicalInspection.objects.create(
                            subsegment=selected_subsegment,
                            defect=defect_for_inspection,
                            status=status_value,
                        )
                    inspection.defect_types.set(
                        DefectType.objects.filter(code__in=selected_defects, is_active=True)
                    )

                    if physical_values["horizontal_alignment"]:
                        horizontal_row = PhysicalInspectionAnalysis.objects.create(
                            inspection=inspection,
                            consideration_type=PhysicalInspectionAnalysis.CONSIDERATION_DESIGN,
                            option=PhysicalInspectionAnalysis.OPTION_HORIZONTAL_ALIGNMENT,
                            option_description=(request.POST.get("horizontal_alignment_description", "") or "").strip(),
                        )
                        PhysicalInspectionCharacteristic.objects.create(
                            analysis=horizontal_row,
                            characteristic=physical_values["horizontal_alignment"],
                            value=physical_values["horizontal_alignment"],
                        )

                    if physical_values["vertical_alignment"]:
                        vertical_row = PhysicalInspectionAnalysis.objects.create(
                            inspection=inspection,
                            consideration_type=PhysicalInspectionAnalysis.CONSIDERATION_DESIGN,
                            option=PhysicalInspectionAnalysis.OPTION_VERTICAL_ALIGNMENT,
                            option_description=(request.POST.get("vertical_alignment_description", "") or "").strip(),
                        )
                        PhysicalInspectionCharacteristic.objects.create(
                            analysis=vertical_row,
                            characteristic=physical_values["vertical_alignment"],
                            value=physical_values["vertical_alignment"],
                        )

                carriage_row = PhysicalInspectionAnalysis.objects.create(
                    inspection=inspection,
                    consideration_type=PhysicalInspectionAnalysis.CONSIDERATION_DESIGN,
                    option=PhysicalInspectionAnalysis.OPTION_CARRIAGE_WAY_CROSS_SECTIONS,
                    option_description=(request.POST.get("carriage_way_description", "") or "").strip(),
                )
                for index in range(1, 9):
                    PhysicalInspectionCharacteristic.objects.create(
                        analysis=carriage_row,
                        characteristic="No of Carriage Ways",
                        value="2",
                        row_index=index,
                    )

                if physical_values["bridges"]:
                    bridges_row = PhysicalInspectionAnalysis.objects.create(
                        inspection=inspection,
                        consideration_type=PhysicalInspectionAnalysis.CONSIDERATION_DESIGN,
                        option=PhysicalInspectionAnalysis.OPTION_BRIDGES,
                        option_description=(request.POST.get("bridges_description", "") or "").strip(),
                    )
                    PhysicalInspectionCharacteristic.objects.create(
                        analysis=bridges_row,
                        characteristic=physical_values["bridges"],
                        value=physical_values["bridges"],
                    )

                for upload in request.FILES.getlist("supporting_files"):
                    Library.objects.create(
                        entry_type=Library.TYPE_PHYSICAL_INSPECTION,
                        file_type=_library_file_type_from_name(upload.name),
                        name=os.path.basename(upload.name),
                        file=upload,
                        defect=defect_for_inspection,
                        physical_inspection=inspection,
                    )
                _sync_defect_status_from_physical(defect_for_inspection, status_value)
                _touch_defect_modified(defect_for_inspection)

                return redirect(
                    reverse("physical_inspection_2" if active_physical_tab == "physical_2" else "physical_inspection")
                )

    physical_inspections = (
        PhysicalInspection.objects.select_related("subsegment", "subsegment__segment")
        .prefetch_related("defect_types", "library_files", "analysis_rows__characteristics")
        .order_by("-id")
    )
    physical_table_statuses = [Defect.WORKFLOW_RCA, Defect.WORKFLOW_PHYSICAL]
    physical_table_defects_qs = (
        Defect.objects.select_related("subsegment", "subsegment__segment", "engineer")
        .filter(workflow_status__in=physical_table_statuses)
        .order_by("-modified", "-id")
    )
    if active_physical_tab == "physical_2":
        physical_draft_defects_qs = Defect.objects.none()
        physical_complete_defects_qs = physical_table_defects_qs.filter(
            workflow_status=Defect.WORKFLOW_PHYSICAL
        )
    else:
        physical_draft_defects_qs = physical_table_defects_qs.filter(
            workflow_status=Defect.WORKFLOW_RCA
        )
        physical_complete_defects_qs = physical_table_defects_qs.filter(
            workflow_status=Defect.WORKFLOW_PHYSICAL
        )
    physical_draft_count = physical_draft_defects_qs.count()
    physical_complete_count = physical_complete_defects_qs.count()
    physical_total = physical_draft_count + physical_complete_count
    defect_label_map = dict(defect_options)
    latest_inspection_by_defect = {}
    for report in physical_inspections:
        if report.defect_id and report.defect_id not in latest_inspection_by_defect:
            latest_inspection_by_defect[report.defect_id] = report
    physical_draft_rows = []
    physical_complete_rows = []
    condition_map = {
        Defect.CONDITION_TOLERABLE: ("Tolerable", "tolerable"),
        Defect.CONDITION_INTOLERABLE: ("Intolerable", "intolerable"),
        Defect.CONDITION_FAILED: ("Intolerable", "intolerable"),
        Defect.CONDITION_BAD: ("No response", "neutral"),
    }
    for defect in physical_draft_defects_qs[:100]:
        report = latest_inspection_by_defect.get(defect.id)
        date_text = f"{timesince(defect.modified).split(',')[0]} ago"

        condition_label, condition_class = condition_map.get(
            defect.condition, ("No response", "neutral")
        )
        engineer_name = "Unassigned"
        if defect.engineer:
            full_name = f"{defect.engineer.first_name} {defect.engineer.last_name}".strip()
            engineer_name = full_name or defect.engineer.get_username()

        physical_draft_rows.append(
            {
                "defect": defect,
                "report": report,
                "status_text": defect.get_workflow_status_display(),
                "condition_label": condition_label,
                "condition_class": condition_class,
                "engineer_name": engineer_name,
                "date_text": date_text,
            }
        )

    for defect in physical_complete_defects_qs[:100]:
        report = latest_inspection_by_defect.get(defect.id)
        condition_label, condition_class = condition_map.get(
            defect.condition, ("No response", "neutral")
        )
        engineer_name = "Unassigned"
        if defect.engineer:
            full_name = f"{defect.engineer.first_name} {defect.engineer.last_name}".strip()
            engineer_name = full_name or defect.engineer.get_username()

        physical_complete_rows.append(
            {
                "defect": defect,
                "report": report,
                "status_text": defect.get_workflow_status_display(),
                "condition_label": condition_label,
                "condition_class": condition_class,
                "engineer_name": engineer_name,
                "date_text": f"{timesince(defect.modified).split(',')[0]} ago",
            }
        )

    current_physical_view = existing_inspection if library_mode == "view" else None
    if (
        library_mode == "view"
        and current_physical_view is None
        and origin_key == "approvals"
        and selected_defect is not None
    ):
        physical_error = "No Physical Inspection report available for this defect."
    elif library_mode == "view" and current_physical_view is None:
        current_physical_view = physical_inspections.filter(
            status=PhysicalInspection.STATUS_COMPLETE
        ).first()

    physical_view_defect_text = ""
    physical_view_cross_sections = []
    physical_supporting_documents = []
    physical_view_segment_status_text = "No response"
    physical_view_segment_status_class = "neutral"
    if current_physical_view:
        option_rows = {row.option: row for row in current_physical_view.analysis_rows.all()}
        horizontal_row = option_rows.get(PhysicalInspectionAnalysis.OPTION_HORIZONTAL_ALIGNMENT)
        if horizontal_row:
            first = horizontal_row.characteristics.first()
            if first:
                physical_values["horizontal_alignment"] = first.value or first.characteristic
        vertical_row = option_rows.get(PhysicalInspectionAnalysis.OPTION_VERTICAL_ALIGNMENT)
        if vertical_row:
            first = vertical_row.characteristics.first()
            if first:
                physical_values["vertical_alignment"] = first.value or first.characteristic
        bridges_row = option_rows.get(PhysicalInspectionAnalysis.OPTION_BRIDGES)
        if bridges_row:
            first = bridges_row.characteristics.first()
            if first:
                physical_values["bridges"] = first.value or first.characteristic
        carriage_row = option_rows.get(
            PhysicalInspectionAnalysis.OPTION_CARRIAGE_WAY_CROSS_SECTIONS
        )
        if carriage_row:
            carriage_items = (
                carriage_row.characteristics.exclude(row_index__isnull=True)
                .order_by("row_index", "id")
            )
            physical_view_cross_sections = [
                {
                    "label": item.characteristic or "No of Carriage Ways",
                    "value": item.value or "Not provided",
                }
                for item in carriage_items[:8]
            ]
        if not physical_view_cross_sections:
            physical_view_cross_sections = [
                {"label": "No of Carriage Ways", "value": "2"} for _ in range(8)
            ]

        selected_codes = list(current_physical_view.defect_types.values_list("code", flat=True))
        selected_labels = [defect_label_map.get(code) for code in selected_codes if code in defect_label_map]
        physical_view_defect_text = ", ".join(label for label in selected_labels if label) or "Not provided"

        attachments = list(
            current_physical_view.library_files.filter(
                entry_type=Library.TYPE_PHYSICAL_INSPECTION
            ).order_by("-created", "-id")
        )
        if attachments:
            physical_supporting_documents = [item.file.name.split("/")[-1] for item in attachments]
        else:
            code = current_physical_view.subsegment.code if current_physical_view.subsegment else "A1LAS2-01"
            physical_supporting_documents = [f"{code} Physical Inspection.jpg", f"{code} Physical Inspection.jpg"]

        segment_status_code = getattr(getattr(current_physical_view.subsegment, "segment", None), "status", "")
        if segment_status_code == "FF5050":
            physical_view_segment_status_text = "Failed"
            physical_view_segment_status_class = "intolerable"
        elif segment_status_code == "FF9966":
            physical_view_segment_status_text = "Intolerable"
            physical_view_segment_status_class = "intolerable"
        elif segment_status_code in {"00CC00", "05700B"}:
            physical_view_segment_status_text = "Tolerable"
            physical_view_segment_status_class = "tolerable"

    return render(
        request,
        "website/engineering_admin.html",
        {
            "active_page": "engineering_admin",
            "active_library_tab": "approvals" if origin_key == "approvals" else active_physical_tab,
            "active_content_tab": active_physical_tab,
            "back_link_text": back_link_text,
            "back_link_href": back_link_href,
            "library_mode": library_mode,
            "physical_success": physical_success,
            "physical_error": physical_error,
            "physical_values": physical_values,
            "current_physical_inspection": existing_inspection,
            "selected_physical_defect": selected_defect,
            "physical_defect_options": defect_options,
            "physical_horizontal_alignment_options": horizontal_alignment_options,
            "physical_vertical_alignment_options": vertical_alignment_options,
            "physical_bridge_options": bridge_options,
            "physical_cross_sections": list(range(8)),
            "physical_total": physical_total,
            "physical_draft_count": physical_draft_count,
            "physical_complete_count": physical_complete_count,
            "physical_draft_rows": physical_draft_rows,
            "physical_complete_rows": physical_complete_rows,
            "current_physical_view": current_physical_view,
            "physical_view_defect_text": physical_view_defect_text,
            "physical_view_cross_sections": physical_view_cross_sections,
            "physical_supporting_documents": physical_supporting_documents,
            "physical_view_segment_status_text": physical_view_segment_status_text,
            "physical_view_segment_status_class": physical_view_segment_status_class,
        },
    )


def engineering_admin_solution_design(request):
    is_history_tab = getattr(getattr(request, "resolver_match", None), "url_name", "") == "library_history"
    active_solution_tab = "history" if is_history_tab else "solution"
    active_solution_route = "library_history" if is_history_tab else "library_solution_design"
    history_engineer = _get_assumed_project_user() if is_history_tab else None
    origin_key = request.GET.get("origin") or request.POST.get("origin") or ""
    default_solution_label = "Archive" if is_history_tab else "Solution Design"
    back_link_text, back_link_href = _resolve_origin_back_link(
        origin_key, default_solution_label, reverse(active_solution_route)
    )
    mode_value = request.GET.get("mode") or request.POST.get("mode")
    if mode_value == "view":
        library_mode = "view"
    else:
        library_mode = "summary"

    selected_defect = None
    selected_defect_id = request.GET.get("defect") or request.POST.get("defect_id")
    if selected_defect_id:
        selected_defect_qs = Defect.objects.select_related("subsegment", "subsegment__segment", "engineer").filter(
            pk=selected_defect_id
        )
        if history_engineer:
            selected_defect_qs = selected_defect_qs.filter(engineer=history_engineer)
        selected_defect = selected_defect_qs.first()

    if request.method == "POST":
        if request.POST.get("command") == "mark_solution_done":
            defect_id = request.POST.get("defect_id")
            defect_qs = Defect.objects.filter(
                pk=defect_id,
                workflow_status=Defect.WORKFLOW_PHYSICAL,
            )
            if history_engineer:
                defect_qs = defect_qs.filter(engineer=history_engineer)
            defect = defect_qs.first()
            if defect:
                defect.workflow_status = Defect.WORKFLOW_SOLUTION
                defect.save(update_fields=["workflow_status", "modified"])
                redirect_url = f"{reverse(active_solution_route)}?sd_done=1"
                if defect_id:
                    redirect_url += f"&mode=view&defect={defect_id}"
                if origin_key:
                    redirect_url += f"&origin={origin_key}"
                return redirect(redirect_url)
            redirect_url = f"{reverse(active_solution_route)}?sd_done=0"
            if defect_id:
                redirect_url += f"&mode=view&defect={defect_id}"
            if origin_key:
                redirect_url += f"&origin={origin_key}"
            return redirect(redirect_url)

        files = request.FILES.getlist("solution_files")
        uploaded_by = _get_assumed_project_user()
        defect_for_upload = None
        defect_id = request.POST.get("defect_id")
        if defect_id:
            defect_for_upload_qs = Defect.objects.filter(pk=defect_id)
            if history_engineer:
                defect_for_upload_qs = defect_for_upload_qs.filter(engineer=history_engineer)
            defect_for_upload = defect_for_upload_qs.first()
        created_count = 0
        for uploaded_file in files:
            if not uploaded_file:
                continue
            Library.objects.create(
                entry_type=Library.TYPE_SOLUTION_DESIGN,
                file_type=_library_file_type_from_name(uploaded_file.name),
                name=os.path.basename(uploaded_file.name),
                file=uploaded_file,
                defect=defect_for_upload,
                uploaded_by=uploaded_by,
            )
            created_count += 1
        if created_count:
            redirect_url = f"{reverse(active_solution_route)}?uploaded={created_count}"
            if defect_for_upload:
                redirect_url += f"&mode=view&defect={defect_for_upload.pk}"
            if origin_key:
                redirect_url += f"&origin={origin_key}"
            return redirect(redirect_url)
        redirect_url = f"{reverse(active_solution_route)}?upload_error=1"
        if defect_for_upload:
            redirect_url += f"&mode=view&defect={defect_for_upload.pk}"
        if origin_key:
            redirect_url += f"&origin={origin_key}"
        return redirect(redirect_url)

    solution_defects_qs = Defect.objects.select_related(
        "subsegment", "subsegment__segment", "engineer"
    ).order_by("-modified", "-id")
    if history_engineer:
        solution_defects_qs = solution_defects_qs.filter(engineer=history_engineer)
    else:
        solution_table_statuses = [Defect.WORKFLOW_PHYSICAL, Defect.WORKFLOW_SOLUTION]
        solution_defects_qs = solution_defects_qs.filter(workflow_status__in=solution_table_statuses)
    solution_defect_ids = list(solution_defects_qs.values_list("id", flat=True))
    solution_file_counts = dict(
        Library.objects.filter(
            entry_type=Library.TYPE_SOLUTION_DESIGN,
            defect_id__in=solution_defect_ids,
        )
        .values("defect_id")
        .annotate(total=Count("id"))
        .values_list("defect_id", "total")
    )
    condition_map = {
        Defect.CONDITION_TOLERABLE: ("Tolerable", "tolerable"),
        Defect.CONDITION_INTOLERABLE: ("Intolerable", "intolerable"),
        Defect.CONDITION_FAILED: ("Intolerable", "intolerable"),
        Defect.CONDITION_BAD: ("No response", "neutral"),
    }
    solution_draft_rows = []
    solution_complete_rows = []
    history_rows = []
    for defect in solution_defects_qs[:100]:
        condition_label, condition_class = condition_map.get(
            defect.condition, ("No response", "neutral")
        )
        engineer_name = "Unassigned"
        if defect.engineer:
            full_name = f"{defect.engineer.first_name} {defect.engineer.last_name}".strip()
            engineer_name = full_name or defect.engineer.get_username()
        row = {
            "defect": defect,
            "status_text": defect.get_workflow_status_display(),
            "condition_label": condition_label,
            "condition_class": condition_class,
            "engineer_name": engineer_name,
            "date_text": f"{timesince(defect.modified).split(',')[0]} ago",
            "files_count": solution_file_counts.get(defect.id, 0),
            "details_url": f"{reverse(active_solution_route)}?mode=view&defect={defect.pk}&origin={active_solution_tab}",
        }
        if defect.workflow_status == Defect.WORKFLOW_SOLUTION:
            solution_complete_rows.append(row)
        else:
            solution_draft_rows.append(row)
        if history_engineer:
            history_rows.append(row)

    solution_view_files = []
    solution_view_status_text = "No response"
    solution_view_status_class = "neutral"
    solution_view_engineer_name = "Unassigned"
    solution_view_segment_code = ""
    if selected_defect:
        solution_view_segment_code = selected_defect.subsegment.code if selected_defect.subsegment else ""
        condition_label, condition_class = condition_map.get(
            selected_defect.condition, ("No response", "neutral")
        )
        solution_view_status_text = condition_label
        solution_view_status_class = condition_class
        if selected_defect.engineer:
            full_name = f"{selected_defect.engineer.first_name} {selected_defect.engineer.last_name}".strip()
            solution_view_engineer_name = full_name or selected_defect.engineer.get_username()
        for item in (
            Library.objects.filter(
                entry_type=Library.TYPE_SOLUTION_DESIGN,
                defect=selected_defect,
            )
            .order_by("-created", "-id")
        ):
            filename = item.name or item.file.name.split("/")[-1]
            ext = os.path.splitext(filename)[1].replace(".", "").upper() or "FILE"
            icon_path = "website/styleguide/icons/sg-file-doc.svg"
            if item.file_type == Library.FILE_TYPE_PDF:
                icon_path = "website/styleguide/icons/sg-file-pdf.svg"
            elif item.file_type == Library.FILE_TYPE_SPREADSHEET:
                icon_path = "website/styleguide/icons/sg-file-xls.svg"
            elif item.file_type == Library.FILE_TYPE_PRESENTATION:
                icon_path = "website/styleguide/icons/sg-file-ppt.svg"
            solution_view_files.append(
                {
                    "id": item.pk,
                    "url": item.file.url if item.file else "#",
                    "name": filename,
                    "ext": ext[:6],
                    "icon_path": icon_path,
                    "date_text": item.created.strftime("%m/%d/%Y"),
                    "size_text": item.file.size if item.file else 0,
                }
            )

    return render(
        request,
        "website/engineering_admin.html",
        {
            "active_page": "engineering_admin",
            "active_library_tab": "approvals" if origin_key == "approvals" else active_solution_tab,
            "active_content_tab": active_solution_tab,
            "back_link_text": back_link_text,
            "back_link_href": back_link_href,
            "origin_key": origin_key,
            "library_mode": library_mode,
            "solution_upload_action_url": reverse(active_solution_route),
            "solution_upload_count": int(request.GET.get("uploaded", "0") or 0),
            "solution_upload_error": request.GET.get("upload_error") == "1",
            "solution_done_success": request.GET.get("sd_done") == "1",
            "solution_total": solution_defects_qs.count(),
            "solution_draft_count": (
                solution_defects_qs.exclude(workflow_status=Defect.WORKFLOW_SOLUTION).count()
                if history_engineer
                else solution_defects_qs.filter(workflow_status=Defect.WORKFLOW_PHYSICAL).count()
            ),
            "solution_complete_count": solution_defects_qs.filter(workflow_status=Defect.WORKFLOW_SOLUTION).count(),
            "solution_draft_rows": solution_draft_rows,
            "solution_complete_rows": solution_complete_rows,
            "history_rows": history_rows,
            "solution_view_defect": selected_defect,
            "solution_view_segment_code": solution_view_segment_code,
            "solution_view_status_text": solution_view_status_text,
            "solution_view_status_class": solution_view_status_class,
            "solution_view_engineer_name": solution_view_engineer_name,
            "solution_view_files": solution_view_files,
            "solution_handoff_message": (
                f"Physical Inspection complete for {selected_defect.subsegment.code}. Upload Solution Design files."
                if selected_defect and request.GET.get("from") == "physical"
                else ""
            ),
        },
    )


def engineering_admin_approvals(request):
    if request.method == "POST":
        is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
        action = request.POST.get("action")
        defect_id = request.POST.get("defect_id")
        defect = Defect.objects.filter(
            pk=defect_id,
            workflow_status=Defect.WORKFLOW_SOLUTION,
        ).first()
        if defect and action in {"approve", "reject", "review"}:
            if action == "approve":
                defect.workflow_status = Defect.WORKFLOW_APPROVED
                defect.save(update_fields=["workflow_status", "modified"])
            elif action == "reject":
                defect.workflow_status = Defect.WORKFLOW_REJECTED
                defect.save(update_fields=["workflow_status", "modified"])
            elif action == "review":
                senior_engineer = (
                    request.user
                    if getattr(request.user, "is_authenticated", False)
                    else _get_assumed_project_user()
                )
                defect.workflow_status = Defect.WORKFLOW_DRAFT
                defect.review = True
                defect.senior_engineer = senior_engineer
                defect.save(update_fields=["workflow_status", "review", "senior_engineer", "modified"])
        if is_ajax:
            pending_count = Defect.objects.filter(workflow_status=Defect.WORKFLOW_SOLUTION).count()
            approved_count = Defect.objects.filter(workflow_status=Defect.WORKFLOW_APPROVED).count()
            rejected_count = Defect.objects.filter(workflow_status=Defect.WORKFLOW_REJECTED).count()
            return JsonResponse(
                {
                    "ok": bool(defect and action in {"approve", "reject", "review"}),
                    "action": action or "",
                    "defect_id": defect_id,
                    "counts": {
                        "total": pending_count + approved_count + rejected_count,
                        "pending": pending_count,
                        "approved": approved_count,
                        "rejected": rejected_count,
                    },
                    "message": (
                        "Approval updated."
                        if defect and action in {"approve", "reject", "review"}
                        else "Approval update failed."
                    ),
                },
                status=200 if defect and action in {"approve", "reject", "review"} else 400,
            )
        return redirect(reverse("library_approvals"))

    def _speed_text(defect):
        subsegment = getattr(defect, "subsegment", None)
        if not subsegment or subsegment.avg_speed is None:
            return "Not available"
        return f"{subsegment.avg_speed}km/hr"

    pending_defects_qs = (
        Defect.objects.select_related("subsegment", "subsegment__segment", "engineer")
        .filter(workflow_status=Defect.WORKFLOW_SOLUTION)
        .order_by("-modified", "-id")
    )
    approved_defects_qs = (
        Defect.objects.select_related("subsegment", "subsegment__segment", "engineer")
        .filter(workflow_status=Defect.WORKFLOW_APPROVED)
        .order_by("-modified", "-id")
    )
    rejected_defects_qs = (
        Defect.objects.select_related("subsegment", "subsegment__segment", "engineer")
        .filter(workflow_status=Defect.WORKFLOW_REJECTED)
        .order_by("-modified", "-id")
    )

    pending_approvals = []
    pending_defect_ids = list(pending_defects_qs.values_list("id", flat=True))
    latest_physical_by_defect = {}
    for inspection in (
        PhysicalInspection.objects.filter(defect_id__in=pending_defect_ids)
        .only("id", "defect_id")
        .order_by("defect_id", "-updated_at", "-id")
    ):
        if inspection.defect_id not in latest_physical_by_defect:
            latest_physical_by_defect[inspection.defect_id] = inspection.id
    for defect in pending_defects_qs:
        engineer_name = "Unassigned"
        if defect.engineer:
            full_name = f"{defect.engineer.first_name} {defect.engineer.last_name}".strip()
            engineer_name = full_name or defect.engineer.get_username()
        rca_url = f"{reverse('engineering_admin')}?mode=view&defect={defect.id}&origin=approvals"
        remedy_url = f"{reverse('library_solution_design')}?mode=view&defect={defect.id}&origin=approvals"
        inspection_id = latest_physical_by_defect.get(defect.id)
        if inspection_id:
            inspection_url = f"{reverse('physical_inspection')}?mode=view&inspection={inspection_id}&origin=approvals"
        else:
            inspection_url = f"{reverse('physical_inspection')}?mode=view&defect={defect.id}&origin=approvals"
        pending_approvals.append(
            {
                "defect_id": defect.id,
                "code": defect.subsegment.code if defect.subsegment else defect.defect_ref or "-",
                "submitted_text": f"Submitted {timesince(defect.modified).split(',')[0]} ago by {engineer_name}",
                "condition": defect.get_condition_display(),
                "speed": _speed_text(defect),
                "rca_detail_url": rca_url,
                "inspection_detail_url": inspection_url,
                "remedy_detail_url": remedy_url,
            }
        )
    approved_plans = [
        {
            "code": defect.subsegment.code if defect.subsegment else defect.defect_ref or "-",
            "age_text": f"{timesince(defect.modified).split(',')[0]} ago",
            "engineer_name": (
                (f"{defect.engineer.first_name} {defect.engineer.last_name}".strip() or defect.engineer.get_username())
                if defect.engineer
                else "Unassigned"
            ),
            "condition": defect.get_condition_display(),
            "speed": _speed_text(defect),
        }
        for defect in approved_defects_qs[:100]
    ]
    rejected_plans = [
        {
            "code": defect.subsegment.code if defect.subsegment else defect.defect_ref or "-",
            "age_text": f"{timesince(defect.modified).split(',')[0]} ago",
            "engineer_name": (
                (f"{defect.engineer.first_name} {defect.engineer.last_name}".strip() or defect.engineer.get_username())
                if defect.engineer
                else "Unassigned"
            ),
            "condition": defect.get_condition_display(),
            "speed": _speed_text(defect),
        }
        for defect in rejected_defects_qs[:100]
    ]

    return render(
        request,
        "website/engineering_admin.html",
        {
            "active_page": "engineering_admin",
            "active_library_tab": "approvals",
            "approvals_total": pending_defects_qs.count() + approved_defects_qs.count() + rejected_defects_qs.count(),
            "approvals_completed": approved_defects_qs.count(),
            "approvals_pending": pending_defects_qs.count(),
            "approvals_rejected": rejected_defects_qs.count(),
            "pending_approvals": pending_approvals,
            "approved_plans": approved_plans,
            "rejected_plans": rejected_plans,
        },
    )

# ---- Road Analysis with page-size selector + robust filter preservation (Point 5) ----



# website/views.py
from decimal import Decimal
from urllib.parse import urlencode
from django.db.models import Sum, Count
from django.core.paginator import Paginator
from django.shortcuts import render, get_object_or_404

from all_roads.models import Segment, Route
from all_roads.models import SubSegment  # <-- NEW import
# assume STATUS_BUCKETS, PAGE_SIZE_DEFAULT, PAGE_SIZE_OPTIONS exist in this module

def road_analysis(request):
    qs = Segment.objects.select_related("route", "start_point", "end_point").all()

    # Query params
    selected_route   = request.GET.get("route") or ""
    selected_state   = (request.GET.get("state") or "").strip()
    if selected_state:
        selected_state = selected_state.upper()
    selected_segment = (request.GET.get("segment") or "").strip()
    show_all         = request.GET.get("show") == "all"

    # Enforce mutual exclusivity (route vs state)
    if show_all:
        selected_route = ""
        selected_state = ""
    elif selected_route:
        selected_state = ""
        qs = qs.filter(route__route=selected_route)
    elif selected_state:
        selected_route = ""
        qs = qs.annotate(_normalized_state=Upper(Trim("state"))).filter(_normalized_state=selected_state)

    # Are we viewing subsegments for a specific segment code?
    is_subsegment_view = False
    sub_qs = None
    parent_segment = None

    if selected_segment:
        # case-insensitive exact match; switch to subsegment mode if found
        parent_segment = get_object_or_404(Segment, code__iexact=selected_segment)
        try:
            refresh_segment_and_subsegments(parent_segment)
            parent_segment.refresh_from_db(fields=["avg_speed", "status", "distance", "travel_time"])
        except Exception as exc:
            logger.warning("Unable to refresh segment %s before rendering subsegments: %s", parent_segment.code, exc)
        sub_qs = SubSegment.objects.filter(segment=parent_segment).order_by("position", "id")
        is_subsegment_view = True

    # Aggregates & counts (only meaningful for the normal segment view)
    if not is_subsegment_view:
        agg = qs.aggregate(total_length=Sum("distance"), total_segments=Count("id"))
        total_length   = (agg["total_length"] or Decimal("0.00")).quantize(Decimal("0.01"))
        total_segments = agg["total_segments"] or 0
        counts = {k: qs.filter(status__in=v["codes"]).count() for k, v in STATUS_BUCKETS.items()}
    else:
        # When viewing subsegments, the top-right panel is a graph, so we don't need totals/metrics.
        total_length = Decimal("0.00")
        total_segments = 0
        counts = {"good": 0, "tolerable": 0, "intolerable": 0, "failed": 0, "no_response": 0}

    # Options for selects
    routes = Route.objects.only("route").order_by("route")
    states = (
        Segment.objects.filter(state__isnull=False)
        .annotate(normalized_state=Upper(Trim("state")))
        .filter(~Q(normalized_state=""))
        .order_by("normalized_state")
        .values_list("normalized_state", flat=True)
        .distinct()
    )

    # Page size
    try:
        page_size = int(request.GET.get("page_size") or PAGE_SIZE_DEFAULT)
        if page_size not in PAGE_SIZE_OPTIONS:
            page_size = PAGE_SIZE_DEFAULT
    except (TypeError, ValueError):
        page_size = PAGE_SIZE_DEFAULT

    # Pagination target: segments OR subsegments
    if is_subsegment_view:
        paginator = Paginator(sub_qs, page_size)
    else:
        qs = qs.order_by("route__route", "index", "code")
        paginator = Paginator(qs, page_size)

    page_obj = paginator.get_page(request.GET.get("page") or 1)
    sn_start = page_obj.start_index() - 1

    # Preserve filters across pagination (strip only 'page')
    qd = request.GET.copy()
    qd.pop("page", None)
    filters_qs = qd.urlencode()

    context = {
        # table data (choose which the template renders)
        "segments": (page_obj.object_list if not is_subsegment_view else []),
        "subsegments": (page_obj.object_list if is_subsegment_view else []),

        "page_obj": page_obj,
        "sn_start": sn_start,

        # filters
        "routes": routes,
        "states": list(states),
        "selected_route": selected_route,
        "selected_state": selected_state,
        "selected_segment": selected_segment,
        "show_all": show_all,

        # metrics (for normal view)
        "total_length": total_length,
        "total_segments": total_segments,
        "counts": counts,
        "filters_qs": filters_qs,

        # page size controls
        "page_size": page_size,
        "page_size_options": PAGE_SIZE_OPTIONS,

        # view mode
        "is_subsegment_view": is_subsegment_view,
        "parent_segment": parent_segment,
        "active_page": "inventory",
    }
    return render(request, "website/road_analysis.html", context)




# ----------------- Helpers and upload pipeline below (unchanged) -----------------

def _in_lat_range(val):
    return Decimal("-90") <= val <= Decimal("90")

def _in_lon_range(val):
    return Decimal("-180") <= val <= Decimal("180")

def _parse_int_or_zero(s):
    try:
        return int(str(s).strip())
    except Exception:
        return 0

def _prime_route_max_index():
    """
    Build a dict: { route_id: current_max_index_int } by casting Segment.index (CharField) to int.
    Non-numeric or blank indexes are treated as 0.
    """
    route_max = defaultdict(int)
    from all_roads.models import Segment  # local import to avoid circulars on module import
    qs = Segment.objects.values("route_id", "index")
    for row in qs:
        route_max[row["route_id"]] = max(route_max[row["route_id"]], _parse_int_or_zero(row["index"]))
    return route_max

def _next_index_for_route(route_obj, cache):
    """
    Return the next two-digit index string per route (01, 02, … 99).
    Uses and updates a mutable cache dict keyed by route_id.
    """
    rid = route_obj.id
    current = cache.get(rid, 0)
    nxt = current + 1
    cache[rid] = nxt
    # zero-pad to length 2 (your model has max_length=2)
    return str(nxt).zfill(2)

try:
    import openpyxl   # for .xlsx
except Exception:
    openpyxl = None

try:
    import xlrd       # for .xls
except Exception:
    xlrd = None

def _road_code_from_route(route_code: str) -> str:
    s = (route_code or "").strip().upper()
    if not s:
        return "F"
    if s[0] == "F":
        return "F"
    if s[0] in ("A", "E"):
        return "A"
    return "F"

REQUIRED_HEADERS = [
    "ROUTE", "SEGMENT CODE"
]

SUBSEG_REQUIRED_HEADERS = ["X_START", "Y_START", "X_END", "Y_END"]

def _to_decimal(s, field_name, rownum, errors):
    try:
        if s is None or s == "":
            return Decimal("0")
        return Decimal(str(s).strip())
    except (InvalidOperation, ValueError):
        errors.append(f"Row {rownum}: invalid decimal for {field_name} = {s!r}")
        return Decimal("0")

def _normalize_headers(headers):
    norm = []
    for h in headers:
        h = (h or "").strip().upper().replace("_", " ")
        norm.append(h)
    return norm


def _normalize_header_token(value):
    value = (value or "").strip().lower()
    value = value.replace("\ufeff", "")
    value = value.replace("_", " ").replace("-", " ")
    value = re.sub(r"[^a-z0-9\s]", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value


SEGMENT_UPLOAD_HEADER_ALIASES = {
    "ROUTE": {"route"},
    "SEGMENT CODE": {"segment code", "segment_code"},
    "STATE": {"state"},
    "SEGMENT NAME": {"segment name"},
    "START NAME": {"start name", "start_name"},
    "START_LAT": {"start lat", "northings"},
    "START_LON": {"start lon", "eastings"},
    "END_NAME": {"end name", "end_name"},
    "END_LAT": {"end lat", "northings 2", "northings2"},
    "END_LON": {"end lon", "eastings 2", "eastings2"},
    "INDEX": {"index"},
}

SUBSEGMENT_UPLOAD_REQUIRED_HEADERS = [
    "SEGMENT",
    "LENGTH",
    "X_START",
    "Y_START",
    "X_END",
    "Y_END",
]

SUBSEGMENT_UPLOAD_HEADER_ALIASES = {
    "SEGMENT": {"segment"},
    "LENGTH": {"length"},
    "X_START": {"x start", "x_start", "x-start", "xstart"},
    "Y_START": {"y start", "y_start", "y-start", "ystart"},
    "X_END": {"x end", "x_end", "x-end", "xend"},
    "Y_END": {"y end", "y_end", "y-end", "yend"},
}

EDIT_SUBSEGMENT_REQUIRED_HEADERS = [
    "SEGMENT",
    "SUBSEGMENT_CODE",
    "LENGTH",
    "X_START",
    "Y_START",
    "X_END",
    "Y_END",
]

EDIT_SUBSEGMENT_HEADER_ALIASES = {
    "SEGMENT": {"segment"},
    "SUBSEGMENT_CODE": {"subsegment code", "subsegment_code", "subsegment", "sub segment code", "code"},
    "LENGTH": {"length"},
    "X_START": {"x start", "x_start", "x-start", "xstart"},
    "Y_START": {"y start", "y_start", "y-start", "ystart"},
    "X_END": {"x end", "x_end", "x-end", "xend"},
    "Y_END": {"y end", "y_end", "y-end", "yend"},
}

def _is_blank_row(cells):
    if not cells:
        return True
    for c in cells:
        if c is None:
            continue
        if isinstance(c, (int, float)):
            return False
        if str(c).strip() != "":
            return False
    return True

def _read_rows(fileobj, filename):
    """
    Yield dicts keyed by canonical segment-upload headers from CSV/XLSX/XLS.
    Required headers: ROUTE, SEGMENT CODE.
    """
    name = (filename or "").lower()

    def _sniff_csv_dialect(data):
        sample = data[:4096]
        try:
            return csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except Exception:
            return csv.excel

    def _build_from_rows(rows):
        if not rows:
            return [], ["The spreadsheet is empty."]

        header_map = {}
        for idx, raw_header in enumerate(rows[0]):
            token = _normalize_header_token(raw_header)
            if not token:
                continue
            for canonical, aliases in SEGMENT_UPLOAD_HEADER_ALIASES.items():
                if token in aliases and canonical not in header_map:
                    header_map[canonical] = idx
                    break

        missing = [h for h in REQUIRED_HEADERS if h not in header_map]
        if missing:
            return [], [f"Missing headers: {', '.join(missing)}"]

        def _cell(r, canonical):
            col = header_map.get(canonical)
            if col is None or col >= len(r):
                return ""
            return r[col]

        out = []
        for i, r in enumerate(rows[1:], start=2):
            if _is_blank_row(r):
                continue
            out.append(
                {
                    "ROUTE": _cell(r, "ROUTE"),
                    "SEGMENT CODE": _cell(r, "SEGMENT CODE"),
                    "STATE": _cell(r, "STATE"),
                    "SEGMENT NAME": _cell(r, "SEGMENT NAME"),
                    "START NAME": _cell(r, "START NAME"),
                    "START_LAT": _cell(r, "START_LAT"),
                    "START_LON": _cell(r, "START_LON"),
                    "END NAME": _cell(r, "END NAME"),
                    "END_LAT": _cell(r, "END_LAT"),
                    "END_LON": _cell(r, "END_LON"),
                    "INDEX": _cell(r, "INDEX"),
                    "_rownum": i,
                }
            )
        return out, []

    if name.endswith(".csv"):
        fileobj.seek(0)
        data = fileobj.read().decode("utf-8", errors="ignore")
        dialect = _sniff_csv_dialect(data)
        rows = list(csv.reader(io.StringIO(data), dialect=dialect))
        return _build_from_rows(rows)

    if name.endswith(".xlsx"):
        if openpyxl is None:
            return [], ["openpyxl not installed (required for .xlsx)"]
        fileobj.seek(0)
        wb = openpyxl.load_workbook(fileobj, data_only=True)
        ws = wb.active
        rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        return _build_from_rows(rows)

    if name.endswith(".xls"):
        if xlrd is None:
            return [], ["xlrd not installed (required for .xls)"]
        fileobj.seek(0)
        book = xlrd.open_workbook(file_contents=fileobj.read())
        sheet = book.sheet_by_index(0)
        rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        return _build_from_rows(rows)

    return [], [f"Unsupported file type: {filename}"]


def _to_decimal_or_none(value):
    try:
        if value is None:
            return None
        parsed = str(value).strip()
        if parsed == "":
            return None
        return Decimal(parsed)
    except (InvalidOperation, TypeError, ValueError):
        return None


def _parse_decimal_required(value, rownum, label, errors, allow_blank_default=None):
    parsed = _to_decimal_or_none(value)
    if parsed is None:
        raw = str(value).strip() if value is not None else ""
        if raw == "" and allow_blank_default is not None:
            return allow_blank_default
        errors.append(f"Row {rownum}: invalid value for {label}.")
        return None
    return parsed


def _extract_segment_index_from_code(segment_code):
    match = re.search(r"(\d{1,2})$", (segment_code or "").strip())
    if not match:
        return None
    return str(int(match.group(1)))


def _process_new_segments_upload(fileobj):
    rows, read_errors = _read_rows(fileobj, getattr(fileobj, "name", ""))
    summary = {
        "created": 0,
        "created_details": [],
        "skipped": 0,
        "skipped_details": [],
        "errors": [],
        "rows_found": len(rows),
    }
    if read_errors:
        summary["errors"].extend(read_errors)
        return summary

    route_cache = {}
    for row in rows:
        rownum = row.get("_rownum")
        route_code = str(row.get("ROUTE") or "").strip().upper()
        segment_code = str(row.get("SEGMENT CODE") or "").strip().upper()

        if not route_code or not segment_code:
            summary["skipped"] += 1
            reason = "required values for Route and Segment Code are missing"
            summary["skipped_details"].append(f"Row {rownum}: {reason}.")
            summary["errors"].append(f"Row {rownum}: {reason}.")
            continue

        derived_index = _extract_segment_index_from_code(segment_code)
        if derived_index is None:
            summary["skipped"] += 1
            reason = "Segment Code must end with one or two digits"
            summary["skipped_details"].append(f"Row {rownum}: {reason}.")
            summary["errors"].append(f"Row {rownum}: {reason}.")
            continue

        if Segment.objects.filter(code__iexact=segment_code).exists():
            summary["skipped"] += 1
            summary["skipped_details"].append(
                f"Row {rownum}: skipped because segment code {segment_code} already exists."
            )
            continue

        route_obj = route_cache.get(route_code)
        if route_obj is None:
            road_code = _road_code_from_route(route_code)
            road_obj, _ = Road.objects.get_or_create(road=road_code)
            route_obj, _ = Route.objects.get_or_create(
                route=route_code,
                defaults={"road": road_obj},
            )
            route_cache[route_code] = route_obj

        segment_kwargs = {
            "route": route_obj,
            "code": segment_code,
            "index": derived_index,
            "state": str(row.get("STATE") or "").strip(),
            "name": str(row.get("SEGMENT NAME") or "").strip(),
        }

        start_lat = _to_decimal_or_none(row.get("START_LAT"))
        start_lon = _to_decimal_or_none(row.get("START_LON"))
        end_lat = _to_decimal_or_none(row.get("END_LAT"))
        end_lon = _to_decimal_or_none(row.get("END_LON"))
        if start_lat is not None:
            segment_kwargs["start_lat"] = start_lat
        if start_lon is not None:
            segment_kwargs["start_lon"] = start_lon
        if end_lat is not None:
            segment_kwargs["end_lat"] = end_lat
        if end_lon is not None:
            segment_kwargs["end_lon"] = end_lon

        try:
            Segment.objects.create(**segment_kwargs)
            summary["created"] += 1
            summary["created_details"].append(
                f"Row {rownum}: created segment {segment_code}."
            )
        except Exception as exc:
            summary["skipped"] += 1
            summary["skipped_details"].append(
                f"Row {rownum}: skipped because segment {segment_code} could not be created."
            )
            summary["errors"].append(f"Row {rownum}: failed to create segment {segment_code}. ({exc})")

    return summary


def _read_new_subsegment_rows(fileobj, filename):
    name = (filename or "").lower()

    def _sniff_csv_dialect(data):
        sample = data[:4096]
        try:
            return csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except Exception:
            return csv.excel

    def _build_from_rows(rows):
        if not rows:
            return [], ["The spreadsheet is empty."]

        header_map = {}
        for idx, raw_header in enumerate(rows[0]):
            token = _normalize_header_token(raw_header)
            if not token:
                continue
            for canonical, aliases in SUBSEGMENT_UPLOAD_HEADER_ALIASES.items():
                if token in aliases and canonical not in header_map:
                    header_map[canonical] = idx
                    break

        missing = [h for h in SUBSEGMENT_UPLOAD_REQUIRED_HEADERS if h not in header_map]
        if missing:
            return [], [f"Missing headers: {', '.join(missing)}"]

        def _cell(r, canonical):
            col = header_map.get(canonical)
            if col is None or col >= len(r):
                return ""
            return r[col]

        out = []
        for i, r in enumerate(rows[1:], start=2):
            if _is_blank_row(r):
                continue
            out.append(
                {
                    "SEGMENT": _cell(r, "SEGMENT"),
                    "LENGTH": _cell(r, "LENGTH"),
                    "X_START": _cell(r, "X_START"),
                    "Y_START": _cell(r, "Y_START"),
                    "X_END": _cell(r, "X_END"),
                    "Y_END": _cell(r, "Y_END"),
                    "_rownum": i,
                }
            )
        return out, []

    if name.endswith(".csv"):
        fileobj.seek(0)
        data = fileobj.read().decode("utf-8", errors="ignore")
        dialect = _sniff_csv_dialect(data)
        rows = list(csv.reader(io.StringIO(data), dialect=dialect))
        return _build_from_rows(rows)

    if name.endswith(".xlsx"):
        if openpyxl is None:
            return [], ["openpyxl not installed (required for .xlsx)"]
        fileobj.seek(0)
        wb = openpyxl.load_workbook(fileobj, data_only=True)
        ws = wb.active
        rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        return _build_from_rows(rows)

    if name.endswith(".xls"):
        if xlrd is None:
            return [], ["xlrd not installed (required for .xls)"]
        fileobj.seek(0)
        book = xlrd.open_workbook(file_contents=fileobj.read())
        sheet = book.sheet_by_index(0)
        rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        return _build_from_rows(rows)

    return [], [f"Unsupported file type: {filename}"]


def _process_new_subsegments_upload(fileobj):
    rows, read_errors = _read_new_subsegment_rows(fileobj, getattr(fileobj, "name", ""))
    summary = {
        "created": 0,
        "created_details": [],
        "skipped": 0,
        "skipped_details": [],
        "errors": [],
        "rows_found": len(rows),
    }
    if read_errors:
        summary["errors"].extend(read_errors)
        return summary

    segment_position_counter = defaultdict(int)
    segment_cache = {}

    for row in rows:
        rownum = row.get("_rownum")
        segment_code = str(row.get("SEGMENT") or "").strip().upper()
        if not segment_code:
            summary["skipped"] += 1
            summary["skipped_details"].append(f"Row {rownum}: skipped because Segment is blank.")
            summary["errors"].append(f"Row {rownum}: Segment is blank.")
            continue

        segment_obj = segment_cache.get(segment_code)
        if segment_obj is None:
            segment_obj = Segment.objects.filter(code__iexact=segment_code).first()
            if segment_obj:
                segment_cache[segment_code] = segment_obj
        if segment_obj is None:
            summary["skipped"] += 1
            summary["skipped_details"].append(
                f"Row {rownum}: skipped because segment {segment_code} was not found."
            )
            summary["errors"].append(f"Row {rownum}: segment {segment_code} was not found.")
            continue

        row_errors = []
        x_start = _parse_decimal_required(row.get("X_START"), rownum, "X_Start", row_errors)
        y_start = _parse_decimal_required(row.get("Y_START"), rownum, "Y_Start", row_errors)
        x_end = _parse_decimal_required(row.get("X_END"), rownum, "X_End", row_errors)
        y_end = _parse_decimal_required(row.get("Y_END"), rownum, "Y_End", row_errors)
        length = _parse_decimal_required(
            row.get("LENGTH"),
            rownum,
            "Length",
            row_errors,
            allow_blank_default=Decimal("0"),
        )

        if not row_errors:
            if not _in_lon_range(x_start):
                row_errors.append(f"Row {rownum}: X_Start must be between -180 and 180.")
            if not _in_lat_range(y_start):
                row_errors.append(f"Row {rownum}: Y_Start must be between -90 and 90.")
            if not _in_lon_range(x_end):
                row_errors.append(f"Row {rownum}: X_End must be between -180 and 180.")
            if not _in_lat_range(y_end):
                row_errors.append(f"Row {rownum}: Y_End must be between -90 and 90.")

        if row_errors:
            summary["skipped"] += 1
            summary["skipped_details"].append(f"Row {rownum}: skipped due to invalid coordinates/length.")
            summary["errors"].extend(row_errors)
            continue

        segment_position_counter[segment_code] += 1
        position = segment_position_counter[segment_code]
        subsegment_code = f"{segment_obj.code}-{position:02d}"

        if SubSegment.objects.filter(code__iexact=subsegment_code).exists():
            summary["skipped"] += 1
            summary["skipped_details"].append(
                f"Row {rownum}: skipped because subsegment code {subsegment_code} already exists."
            )
            continue

        if SubSegment.objects.filter(segment=segment_obj, position=position).exists():
            summary["skipped"] += 1
            summary["skipped_details"].append(
                f"Row {rownum}: skipped because position {position} already exists for segment {segment_obj.code}."
            )
            continue

        try:
            SubSegment.objects.create(
                segment=segment_obj,
                position=position,
                code=subsegment_code,
                start_lon=x_start,
                start_lat=y_start,
                end_lon=x_end,
                end_lat=y_end,
                distance=length,
            )
            summary["created"] += 1
            summary["created_details"].append(
                f"Row {rownum}: created subsegment {subsegment_code}."
            )
        except Exception as exc:
            summary["skipped"] += 1
            summary["skipped_details"].append(
                f"Row {rownum}: skipped because subsegment {subsegment_code} could not be created."
            )
            summary["errors"].append(f"Row {rownum}: failed to create subsegment {subsegment_code}. ({exc})")

    return summary


def _process_edit_segments_upload(fileobj):
    rows, read_errors = _read_rows(fileobj, getattr(fileobj, "name", ""))
    summary = {
        "updated": 0,
        "updated_details": [],
        "skipped": 0,
        "skipped_details": [],
        "errors": [],
        "rows_found": len(rows),
    }
    if read_errors:
        summary["errors"].extend(read_errors)
        return summary

    for row in rows:
        rownum = row.get("_rownum")
        route_code = str(row.get("ROUTE") or "").strip().upper()
        segment_code = str(row.get("SEGMENT CODE") or "").strip().upper()
        if not route_code or not segment_code:
            summary["skipped"] += 1
            summary["errors"].append(f"Row {rownum}: Route and Segment Code are required.")
            summary["skipped_details"].append(f"Row {rownum}: skipped due to missing Route/Segment Code.")
            continue

        segment = Segment.objects.select_related("route").filter(code__iexact=segment_code).first()
        if not segment:
            summary["skipped"] += 1
            summary["errors"].append(f"Row {rownum}: segment {segment_code} was not found.")
            summary["skipped_details"].append(f"Row {rownum}: skipped because segment {segment_code} does not exist.")
            continue

        current_route = (segment.route.route if segment.route_id and segment.route else "").strip().upper()
        if current_route != route_code:
            summary["skipped"] += 1
            summary["errors"].append(
                f"Row {rownum}: route mismatch for {segment_code}. Existing route is {current_route}, file has {route_code}."
            )
            summary["skipped_details"].append(f"Row {rownum}: skipped because route change is not allowed.")
            continue

        updates = {}

        state_raw = str(row.get("STATE") or "")
        if state_raw.strip():
            updates["state"] = state_raw.strip()

        name_raw = str(row.get("SEGMENT NAME") or "")
        if name_raw.strip():
            updates["name"] = name_raw.strip()

        index_raw = str(row.get("INDEX") or "")
        if index_raw.strip():
            updates["index"] = index_raw.strip()

        decimal_errors = []
        for key, field in (
            ("START_LAT", "start_lat"),
            ("START_LON", "start_lon"),
            ("END_LAT", "end_lat"),
            ("END_LON", "end_lon"),
        ):
            raw = row.get(key)
            raw_text = str(raw).strip() if raw is not None else ""
            if raw_text == "":
                continue
            parsed = _to_decimal_or_none(raw)
            if parsed is None:
                decimal_errors.append(f"Row {rownum}: invalid value for {key}.")
            else:
                updates[field] = parsed

        if decimal_errors:
            summary["skipped"] += 1
            summary["errors"].extend(decimal_errors)
            summary["skipped_details"].append(f"Row {rownum}: skipped due to invalid numeric data.")
            continue

        if not updates:
            summary["skipped"] += 1
            summary["skipped_details"].append(f"Row {rownum}: skipped because no editable values were provided.")
            continue

        try:
            for field, value in updates.items():
                setattr(segment, field, value)
            segment.save(update_fields=list(updates.keys()))
            summary["updated"] += 1
            summary["updated_details"].append(f"Row {rownum}: updated segment {segment_code}.")
        except Exception as exc:
            summary["skipped"] += 1
            summary["errors"].append(f"Row {rownum}: failed to update segment {segment_code}. ({exc})")
            summary["skipped_details"].append(f"Row {rownum}: skipped because update failed for {segment_code}.")

    return summary


def _read_edit_subsegment_rows(fileobj, filename):
    name = (filename or "").lower()

    def _sniff_csv_dialect(data):
        sample = data[:4096]
        try:
            return csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except Exception:
            return csv.excel

    def _build_from_rows(rows):
        if not rows:
            return [], ["The spreadsheet is empty."]

        header_map = {}
        for idx, raw_header in enumerate(rows[0]):
            token = _normalize_header_token(raw_header)
            if not token:
                continue
            for canonical, aliases in EDIT_SUBSEGMENT_HEADER_ALIASES.items():
                if token in aliases and canonical not in header_map:
                    header_map[canonical] = idx
                    break

        missing = [h for h in EDIT_SUBSEGMENT_REQUIRED_HEADERS if h not in header_map]
        if missing:
            return [], [f"Missing headers: {', '.join(missing)}"]

        def _cell(r, canonical):
            col = header_map.get(canonical)
            if col is None or col >= len(r):
                return ""
            return r[col]

        out = []
        for i, r in enumerate(rows[1:], start=2):
            if _is_blank_row(r):
                continue
            out.append(
                {
                    "SEGMENT": _cell(r, "SEGMENT"),
                    "SUBSEGMENT_CODE": _cell(r, "SUBSEGMENT_CODE"),
                    "LENGTH": _cell(r, "LENGTH"),
                    "X_START": _cell(r, "X_START"),
                    "Y_START": _cell(r, "Y_START"),
                    "X_END": _cell(r, "X_END"),
                    "Y_END": _cell(r, "Y_END"),
                    "_rownum": i,
                }
            )
        return out, []

    if name.endswith(".csv"):
        fileobj.seek(0)
        data = fileobj.read().decode("utf-8", errors="ignore")
        dialect = _sniff_csv_dialect(data)
        rows = list(csv.reader(io.StringIO(data), dialect=dialect))
        return _build_from_rows(rows)

    if name.endswith(".xlsx"):
        if openpyxl is None:
            return [], ["openpyxl not installed (required for .xlsx)"]
        fileobj.seek(0)
        wb = openpyxl.load_workbook(fileobj, data_only=True)
        ws = wb.active
        rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        return _build_from_rows(rows)

    if name.endswith(".xls"):
        if xlrd is None:
            return [], ["xlrd not installed (required for .xls)"]
        fileobj.seek(0)
        book = xlrd.open_workbook(file_contents=fileobj.read())
        sheet = book.sheet_by_index(0)
        rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        return _build_from_rows(rows)

    return [], [f"Unsupported file type: {filename}"]


def _process_edit_subsegments_upload(fileobj):
    rows, read_errors = _read_edit_subsegment_rows(fileobj, getattr(fileobj, "name", ""))
    summary = {
        "updated": 0,
        "updated_details": [],
        "skipped": 0,
        "skipped_details": [],
        "errors": [],
        "rows_found": len(rows),
    }
    if read_errors:
        summary["errors"].extend(read_errors)
        return summary

    for row in rows:
        rownum = row.get("_rownum")
        segment_code = str(row.get("SEGMENT") or "").strip().upper()
        subsegment_code = str(row.get("SUBSEGMENT_CODE") or "").strip().upper()
        if not segment_code or not subsegment_code:
            summary["skipped"] += 1
            summary["errors"].append(f"Row {rownum}: Segment and Subsegment code are required.")
            summary["skipped_details"].append(f"Row {rownum}: skipped due to missing Segment/Subsegment code.")
            continue

        subsegment = SubSegment.objects.select_related("segment").filter(code__iexact=subsegment_code).first()
        if not subsegment:
            summary["skipped"] += 1
            summary["errors"].append(f"Row {rownum}: subsegment {subsegment_code} was not found.")
            summary["skipped_details"].append(f"Row {rownum}: skipped because subsegment {subsegment_code} does not exist.")
            continue

        if not subsegment.segment_id or (subsegment.segment.code or "").strip().upper() != segment_code:
            summary["skipped"] += 1
            summary["errors"].append(
                f"Row {rownum}: segment mismatch for {subsegment_code}. Existing segment is {subsegment.segment.code if subsegment.segment_id else '-'}, file has {segment_code}."
            )
            summary["skipped_details"].append(f"Row {rownum}: skipped because segment/subsegment mismatch.")
            continue

        updates = {}
        row_errors = []

        for key, field, is_lon, is_lat in (
            ("X_START", "start_lon", True, False),
            ("Y_START", "start_lat", False, True),
            ("X_END", "end_lon", True, False),
            ("Y_END", "end_lat", False, True),
        ):
            raw = row.get(key)
            raw_text = str(raw).strip() if raw is not None else ""
            if raw_text == "":
                continue
            parsed = _to_decimal_or_none(raw)
            if parsed is None:
                row_errors.append(f"Row {rownum}: invalid value for {key}.")
                continue
            if is_lon and not _in_lon_range(parsed):
                row_errors.append(f"Row {rownum}: {key} must be between -180 and 180.")
                continue
            if is_lat and not _in_lat_range(parsed):
                row_errors.append(f"Row {rownum}: {key} must be between -90 and 90.")
                continue
            updates[field] = parsed

        length_raw = row.get("LENGTH")
        length_text = str(length_raw).strip() if length_raw is not None else ""
        if length_text != "":
            parsed_length = _to_decimal_or_none(length_raw)
            if parsed_length is None:
                row_errors.append(f"Row {rownum}: invalid value for LENGTH.")
            else:
                updates["distance"] = parsed_length

        if row_errors:
            summary["skipped"] += 1
            summary["errors"].extend(row_errors)
            summary["skipped_details"].append(f"Row {rownum}: skipped due to invalid data.")
            continue

        if not updates:
            summary["skipped"] += 1
            summary["skipped_details"].append(f"Row {rownum}: skipped because no editable values were provided.")
            continue

        try:
            for field, value in updates.items():
                setattr(subsegment, field, value)
            subsegment.save(update_fields=list(updates.keys()))
            summary["updated"] += 1
            summary["updated_details"].append(f"Row {rownum}: updated subsegment {subsegment_code}.")
        except Exception as exc:
            summary["skipped"] += 1
            summary["errors"].append(f"Row {rownum}: failed to update subsegment {subsegment_code}. ({exc})")
            summary["skipped_details"].append(f"Row {rownum}: skipped because update failed for {subsegment_code}.")

    return summary


def library_upload_dispatch(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "POST method required."}, status=405)

    upload_option = (request.POST.get("upload_option") or "").strip()
    upload_file = request.FILES.get("upload_file")
    if not upload_option:
        return JsonResponse({"ok": False, "message": "Select an upload option."}, status=400)
    if not upload_file:
        return JsonResponse({"ok": False, "message": "Select a file to upload."}, status=400)

    ext = os.path.splitext((upload_file.name or "").lower())[1]
    if ext not in {".xls", ".xlsx", ".csv"}:
        return JsonResponse(
            {"ok": False, "message": "Unsupported file type. Allowed: .xls, .xlsx, .csv"},
            status=400,
        )

    if upload_option == "new_segments":
        summary = _process_new_segments_upload(upload_file)
        return JsonResponse(
            {
                "ok": True,
                "message": "Upload processed.",
                "option": upload_option,
                "summary": summary,
            }
        )
    if upload_option == "new_1km_subsegments":
        summary = _process_new_subsegments_upload(upload_file)
        return JsonResponse(
            {
                "ok": True,
                "message": "Upload processed.",
                "option": upload_option,
                "summary": summary,
            }
        )
    if upload_option == "edit_segments":
        summary = _process_edit_segments_upload(upload_file)
        return JsonResponse(
            {
                "ok": True,
                "message": "Upload processed.",
                "option": upload_option,
                "summary": summary,
            }
        )
    if upload_option == "edit_subsegments":
        summary = _process_edit_subsegments_upload(upload_file)
        return JsonResponse(
            {
                "ok": True,
                "message": "Upload processed.",
                "option": upload_option,
                "summary": summary,
            }
        )

    return JsonResponse(
        {
            "ok": False,
            "message": "This upload option is not wired yet.",
            "option": upload_option,
        },
        status=400,
    )

def _read_subsegment_rows(fileobj, filename):
    """
    Extract the coordinate columns needed for sub-segment uploads.
    """
    name = (filename or "").lower()
    required_norm = _normalize_headers(SUBSEG_REQUIRED_HEADERS)
    norm_to_canon = dict(zip(required_norm, SUBSEG_REQUIRED_HEADERS))

    def _build_from_rows(rows):
        if not rows:
            return [], ["The spreadsheet is empty."]
        headers = _normalize_headers(rows[0])
        idx = {norm_to_canon[h]: headers.index(h) for h in required_norm if h in headers}
        missing = [norm_to_canon[h] for h in required_norm if h not in headers]
        if missing:
            missing_titles = ", ".join(missing).lower()
            return [], [f"Missing column headers: {missing_titles}."]
        out = []
        for i, r in enumerate(rows[1:], start=2):
            if _is_blank_row(r):
                continue
            def cell(h):
                j = idx[h]
                return r[j] if j < len(r or []) else ""
            out.append({
                "X_START": cell("X_START"),
                "Y_START": cell("Y_START"),
                "X_END": cell("X_END"),
                "Y_END": cell("Y_END"),
                "_rownum": i,
            })
        return out, []

    if name.endswith(".csv"):
        fileobj.seek(0)
        data = fileobj.read().decode("utf-8", errors="ignore")
        rows = list(csv.reader(io.StringIO(data)))
        return _build_from_rows(rows)
    elif name.endswith(".xlsx"):
        if openpyxl is None:
            return [], ["Excel support is unavailable on this server (.xlsx)."]
        fileobj.seek(0)
        wb = openpyxl.load_workbook(fileobj, data_only=True)
        ws = wb.active
        rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        return _build_from_rows(rows)
    elif name.endswith(".xls"):
        if xlrd is None:
            return [], ["Excel support is unavailable on this server (.xls)."]
        fileobj.seek(0)
        book = xlrd.open_workbook(file_contents=fileobj.read())
        sheet = book.sheet_by_index(0)
        rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        return _build_from_rows(rows)
    elif name.endswith(".numbers"):
        return [], ["Apple Numbers files must be exported to CSV or Excel before uploading."]
    else:
        return [], [f"Unsupported file type: {filename}"]

def _process_subsegment_upload(segment_code, start_row, end_row, fileobj):
    errors = []
    seg_code = (segment_code or "").strip()
    if not seg_code:
        return {"created": 0, "replaced": 0, "errors": ["Please select a segment."]}

    try:
        segment_obj = Segment.objects.get(code__iexact=seg_code)
    except Segment.DoesNotExist:
        return {
            "created": 0,
            "replaced": 0,
            "errors": [f"The segment “{seg_code}” was not found. Please check the code and try again."]
        }

    rows, header_errors = _read_subsegment_rows(fileobj, fileobj.name)
    if header_errors:
        return {"created": 0, "replaced": 0, "errors": header_errors}

    if not rows:
        return {"created": 0, "replaced": 0, "errors": ["No sub-segments were found in the spreadsheet."]}

    row_lookup = {row["_rownum"]: row for row in rows}
    selected_rows = []
    missing_rows = []
    for row_num in range(start_row, end_row + 1):
        data = row_lookup.get(row_num)
        if data:
            selected_rows.append(data)
        else:
            missing_rows.append(row_num)

    if missing_rows:
        pretty_rows = ", ".join(str(num) for num in missing_rows)
        return {
            "created": 0,
            "replaced": 0,
            "errors": [f"Rows {pretty_rows} have no data. Update the sheet or adjust the range."]
        }

    cleaned_rows = []
    for row in selected_rows:
        rownum = row.get("_rownum")
        coord_errors_before = len(errors)
        start_lon = _to_decimal(row["X_START"], "x_start", rownum, errors)
        start_lat = _to_decimal(row["Y_START"], "y_start", rownum, errors)
        end_lon   = _to_decimal(row["X_END"], "x_end", rownum, errors)
        end_lat   = _to_decimal(row["Y_END"], "y_end", rownum, errors)
        has_new_error = len(errors) > coord_errors_before

        coord_bad = False
        if not _in_lon_range(start_lon):
            errors.append(f"Row {rownum}: x_start must be between -180 and 180.")
            coord_bad = True
        if not _in_lat_range(start_lat):
            errors.append(f"Row {rownum}: y_start must be between -90 and 90.")
            coord_bad = True
        if not _in_lon_range(end_lon):
            errors.append(f"Row {rownum}: x_end must be between -180 and 180.")
            coord_bad = True
        if not _in_lat_range(end_lat):
            errors.append(f"Row {rownum}: y_end must be between -90 and 90.")
            coord_bad = True

        if has_new_error or coord_bad:
            continue

        cleaned_rows.append({
            "start_lat": start_lat,
            "start_lon": start_lon,
            "end_lat": end_lat,
            "end_lon": end_lon,
            "_rownum": rownum,
        })

    if errors:
        return {"created": 0, "replaced": 0, "errors": errors}

    with transaction.atomic():
        replaced = SubSegment.objects.filter(segment=segment_obj).count()
        SubSegment.objects.filter(segment=segment_obj).delete()
        objs = []
        for position, payload in enumerate(cleaned_rows, start=1):
            objs.append(SubSegment(
                segment=segment_obj,
                position=position,
                code=f"{segment_obj.code}-{position:02d}",
                start_lat=payload["start_lat"],
                start_lon=payload["start_lon"],
                end_lat=payload["end_lat"],
                end_lon=payload["end_lon"],
            ))
        SubSegment.objects.bulk_create(objs)

    return {"created": len(cleaned_rows), "replaced": replaced, "errors": []}

def segment_code_search(request):
    """
    Return top matching segment codes for typeahead.
    GET /segments/search/?q=AB -> { "results": ["AB01", "AB02", ...] }
    """
    q = (request.GET.get("q") or "").strip()
    if len(q) < 2:
        return JsonResponse({"results": []})
    qs = Segment.objects.order_by("code")
    qs = qs.filter(code__icontains=q)
    codes = list(qs.values_list("code", flat=True).distinct()[:20])  # cap results
    return JsonResponse({"results": codes})


def road_inventory_route_details(request):
    route_code = (request.GET.get("route") or "").strip()
    if not route_code:
        return JsonResponse({"ok": False, "message": "Route is required."}, status=400)

    route_obj = Route.objects.filter(route=route_code).only("route", "details").first()
    if not route_obj:
        return JsonResponse({"ok": False, "message": f'Route "{route_code}" was not found.'}, status=404)

    seg_qs = Segment.objects.filter(route__route=route_code).order_by("index", "code")
    first_segment = seg_qs.select_related("start_point").first()
    last_segment = seg_qs.select_related("end_point").last()
    summary_start_point = ""
    summary_end_point = ""
    if first_segment:
        start_name = (
            first_segment.start_point.name
            if first_segment.start_point_id and first_segment.start_point and first_segment.start_point.name
            else (first_segment.name or first_segment.state or "")
        )
        summary_start_point = _format_segment_point_display(
            start_name,
            first_segment.start_lat,
            first_segment.start_lon,
        )
    if last_segment:
        end_name = (
            last_segment.end_point.name
            if last_segment.end_point_id and last_segment.end_point and last_segment.end_point.name
            else (last_segment.name or last_segment.state or "")
        )
        summary_end_point = _format_segment_point_display(
            end_name,
            last_segment.end_lat,
            last_segment.end_lon,
        )
    route_total_length = seg_qs.aggregate(total_length=Sum("distance")).get("total_length")
    rows = [
        {
            "code": seg.code,
            "state": seg.state or "-",
            "name": seg.name or "-",
            "start_point": seg.start_point.name if seg.start_point_id and seg.start_point and seg.start_point.name else "-",
            "end_point": seg.end_point.name if seg.end_point_id and seg.end_point and seg.end_point.name else "-",
            "distance": str(seg.distance) if seg.distance is not None else "-",
            "settlement_type": seg.settlement_type or "-",
            "carriages": seg.carriages if seg.carriages is not None else "-",
            "lanes": seg.lanes if seg.lanes is not None else "-",
            "pavement_type": seg.pavement_type or "-",
            "junctions": seg.junctions if seg.junctions is not None else "-",
            "culverts": seg.culverts if seg.culverts is not None else "-",
            "bridges": seg.bridges if seg.bridges is not None else "-",
        }
        for seg in seg_qs.select_related("start_point", "end_point")
    ]

    return JsonResponse(
        {
            "ok": True,
            "summary": {
                "route": route_obj.route,
                "length": _format_km_total(route_total_length),
                "start_point": summary_start_point,
                "end_point": summary_end_point,
                "passes_through": route_obj.details or "",
                "number_of_segments": len(rows),
            },
            "segments": rows,
        }
    )


def road_condition_subsegments(request):
    """
    AJAX endpoint for road_condition:
    Case-insensitive segment lookup by code, then return its subsegments.
    """
    segment_code = (request.GET.get("segment") or "").strip()
    if not segment_code:
        return JsonResponse(
            {"ok": False, "message": "Please select a segment code.", "rows": []},
            status=400,
        )

    segment = Segment.objects.filter(code__iexact=segment_code).first()
    if not segment:
        return JsonResponse(
            {"ok": False, "message": f'Segment "{segment_code}" was not found.', "rows": []},
            status=404,
        )

    refresh_requested = (request.GET.get("refresh") or "").strip() in {"1", "true", "yes"}
    refresh_meta = {"attempted": False, "failed": False, "updated": 0, "failed_count": 0, "total": 0}
    refresh_warning = ""

    if refresh_requested:
        refresh_meta["attempted"] = True
        try:
            refresh_result = refresh_subsegments_from_google(
                SubSegment.objects.filter(segment=segment),
                sleep_between=0.0,
            )
            refresh_meta.update(
                {
                    "updated": int(refresh_result.get("updated", 0) or 0),
                    "failed_count": int(refresh_result.get("failed", 0) or 0),
                    "total": int(refresh_result.get("total", 0) or 0),
                }
            )
        except Exception:
            refresh_meta["failed"] = True
            refresh_warning = " Refresh failed; showing saved data."

    sub_qs = SubSegment.objects.filter(segment=segment).order_by("position", "id")
    weighted_distance_total = Decimal("0")
    weighted_speed_total = Decimal("0")
    for row in sub_qs:
        distance_value = row.distance or Decimal("0")
        speed_value = row.avg_speed or Decimal("0")
        if distance_value and distance_value > 0:
            weighted_distance_total += distance_value
            weighted_speed_total += distance_value * speed_value

    if weighted_distance_total > 0:
        segment_avg_speed = float((weighted_speed_total / weighted_distance_total).quantize(Decimal("0.1")))
    else:
        segment_avg_speed = 0.0

    rows = [
        {
            "segment_code": segment.code,
            "code": row.code or f"{segment.code}-{row.position:02d}",
            "position": row.position,
            "start_lat": str(row.start_lat),
            "start_lon": str(row.start_lon),
            "distance": str(row.distance),
            "avg_speed": float(row.avg_speed or 0),
            "status": row.status or "666699",
        }
        for row in sub_qs
    ]

    if not rows:
        return JsonResponse(
            {
                "ok": True,
                "message": f'Segment "{segment.code}" has no subsegments.{refresh_warning}',
                "rows": [],
                "segment_avg_speed": segment_avg_speed,
                "segment_status": segment.status or "666699",
                "refresh": refresh_meta,
            }
        )

    if refresh_meta["attempted"] and not refresh_meta["failed"]:
        message = (
            f'Refreshed {refresh_meta["updated"]}/{refresh_meta["total"]} subsegments '
            f'for "{segment.code}" (failed: {refresh_meta["failed_count"]}).'
        )
    else:
        message = f'Loaded {len(rows)} subsegments for "{segment.code}".{refresh_warning}'

    return JsonResponse(
        {
            "ok": True,
            "message": message,
            "rows": rows,
            "segment_avg_speed": segment_avg_speed,
            "segment_status": segment.status or "666699",
            "refresh": refresh_meta,
        }
    )
